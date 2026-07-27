from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Literal
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.models import StockBatch
from app.services.inventory import stock_summary


ReportLanguage = Literal["en", "ru", "uz"]

REPORT_TEXT = {
    "en": {
        "title": "Material Inventory Report",
        "subtitle": "Positive on-hand material stock grouped by material",
        "generated": "Generated",
        "supplier": "Supplier",
        "no_supplier": "No supplier",
        "number": "No.",
        "material": "Material name",
        "sku": "SKU",
        "batches": "Batch rows",
        "count": "Rolls / pieces",
        "kg": "Total kg",
        "grand_total": "Grand total",
        "page": "Page",
    },
    "ru": {
        "title": "Отчёт по складу материалов",
        "subtitle": "Положительные остатки материалов, сгруппированные по материалу",
        "generated": "Сформировано",
        "supplier": "Поставщик",
        "no_supplier": "Без поставщика",
        "number": "№",
        "material": "Наименование материала",
        "sku": "SKU",
        "batches": "Партий",
        "count": "Рулоны / штуки",
        "kg": "Всего кг",
        "grand_total": "Общий итог",
        "page": "Страница",
    },
    "uz": {
        "title": "Materiallar ombori hisoboti",
        "subtitle": "Material bo'yicha guruhlangan musbat ombor qoldiqlari",
        "generated": "Yaratildi",
        "supplier": "Yetkazib beruvchi",
        "no_supplier": "Yetkazib beruvchi ko'rsatilmagan",
        "number": "№",
        "material": "Material nomi",
        "sku": "SKU",
        "batches": "Partiyalar",
        "count": "Rulon / dona",
        "kg": "Jami kg",
        "grand_total": "Umumiy jami",
        "page": "Sahifa",
    },
}


def material_inventory_report_rows(
    db: Session,
    *,
    supplier_id: int | None = None,
    supplier_unassigned: bool = False,
) -> list[dict]:
    stock_rows = stock_summary(
        db,
        group="materials",
        supplier_id=supplier_id,
        supplier_unassigned=supplier_unassigned,
        positive_only=True,
    )
    item_ids = [int(row["item_id"]) for row in stock_rows]
    batch_totals: dict[int, tuple[int, int]] = {}
    if item_ids:
        query = (
            db.query(
                StockBatch.item_id,
                func.count(StockBatch.id),
                func.coalesce(func.sum(StockBatch.piece_count), 0),
            )
            .filter(StockBatch.item_id.in_(item_ids), StockBatch.quantity > 0)
        )
        if supplier_id is not None:
            query = query.filter(StockBatch.supplier_id == supplier_id)
        elif supplier_unassigned:
            query = query.filter(StockBatch.supplier_id.is_(None))
        rows = query.group_by(StockBatch.item_id).all()
        batch_totals = {
            int(item_id): (int(batch_count or 0), int(piece_count or 0))
            for item_id, batch_count, piece_count in rows
        }

    report_rows = []
    for row in stock_rows:
        quantity = float(row.get("quantity") or 0)
        if quantity <= 0:
            continue
        item_id = int(row["item_id"])
        batch_count, piece_count = batch_totals.get(item_id, (0, 0))
        report_rows.append({
            "item_id": item_id,
            "material_name": str(row.get("name") or ""),
            "sku": str(row.get("sku") or ""),
            "batch_count": batch_count,
            "piece_count": piece_count,
            "total_kg": quantity if str(row.get("unit") or "").lower() == "kg" else 0.0,
        })
    return sorted(report_rows, key=lambda row: (row["material_name"].casefold(), row["sku"].casefold()))


def material_inventory_supplier_scope_label(
    lang: ReportLanguage,
    supplier_name: str | None,
    supplier_unassigned: bool,
) -> str | None:
    if not supplier_name and not supplier_unassigned:
        return None
    text = REPORT_TEXT[lang]
    supplier_value = text["no_supplier"] if supplier_unassigned else str(supplier_name or "")
    return f'{text["supplier"]}: {supplier_value}'


def build_material_inventory_xlsx(
    rows: list[dict],
    generated_label: str,
    lang: ReportLanguage,
    *,
    scope_label: str | None = None,
) -> bytes:
    text = REPORT_TEXT[lang]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Material Inventory"
    workbook.properties.title = text["title"]
    workbook.properties.subject = text["subtitle"]

    dark_fill = PatternFill("solid", fgColor="1F1C17")
    light_fill = PatternFill("solid", fgColor="EEEAE0")
    white_font = Font(color="FFFFFF", bold=True)
    title_font = Font(size=16, bold=True, color="1F1C17")
    thin_border = Border(bottom=Side(style="thin", color="D8D1C0"))

    sheet.merge_cells("A1:F1")
    sheet["A1"] = text["title"]
    sheet["A1"].font = title_font
    sheet["A1"].alignment = Alignment(horizontal="left")
    sheet.merge_cells("A2:F2")
    sheet["A2"] = text["subtitle"]
    sheet["A2"].font = Font(size=10, color="625B4B")
    sheet.merge_cells("A3:F3")
    sheet["A3"] = " | ".join(
        value
        for value in (f'{text["generated"]}: {generated_label}', scope_label)
        if value
    )
    sheet["A3"].font = Font(size=9, color="7A725F")

    header_row = 5
    headers = [
        text["number"],
        text["material"],
        text["sku"],
        text["batches"],
        text["count"],
        text["kg"],
    ]
    for column, value in enumerate(headers, 1):
        cell = sheet.cell(header_row, column, value)
        cell.fill = dark_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    first_data_row = header_row + 1
    for index, row in enumerate(rows, 1):
        excel_row = header_row + index
        values = [
            index,
            row["material_name"],
            row["sku"],
            int(row["batch_count"]),
            int(row["piece_count"]),
            float(row["total_kg"]),
        ]
        for column, value in enumerate(values, 1):
            cell = sheet.cell(excel_row, column, value)
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal="right" if column in {1, 4, 5, 6} else "left",
                vertical="center",
            )
        sheet.cell(excel_row, 6).number_format = '#,##0.00'

    total_row = header_row + len(rows) + 1
    sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
    sheet.cell(total_row, 1, text["grand_total"])
    sheet.cell(total_row, 1).alignment = Alignment(horizontal="right")
    if rows:
        sheet.cell(total_row, 4, f"=SUM(D{first_data_row}:D{total_row - 1})")
        sheet.cell(total_row, 5, f"=SUM(E{first_data_row}:E{total_row - 1})")
        sheet.cell(total_row, 6, f"=SUM(F{first_data_row}:F{total_row - 1})")
    else:
        for column in (4, 5, 6):
            sheet.cell(total_row, column, 0)
    for column in range(1, 7):
        cell = sheet.cell(total_row, column)
        cell.fill = light_fill
        cell.font = Font(bold=True, color="1F1C17")
        cell.border = Border(top=Side(style="medium", color="1F1C17"))
    sheet.cell(total_row, 6).number_format = '#,##0.00'

    widths = {"A": 8, "B": 46, "C": 24, "D": 14, "E": 18, "F": 16}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.row_dimensions[header_row].height = 24
    sheet.freeze_panes = "A6"
    sheet.auto_filter.ref = f"A{header_row}:F{max(header_row, total_row - 1)}"
    sheet.sheet_view.showGridLines = False
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.4
    sheet.page_margins.bottom = 0.4
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _register_report_fonts() -> tuple[str, str]:
    regular_candidates = [
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
        Path("C:/Windows/Fonts/arial.ttf"),
    ]
    bold_candidates = [
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
        Path("C:/Windows/Fonts/arialbd.ttf"),
    ]
    regular_path = next((path for path in regular_candidates if path.exists()), None)
    bold_path = next((path for path in bold_candidates if path.exists()), None)
    if not regular_path or not bold_path:
        raise RuntimeError("A Unicode TrueType font is required for inventory PDF reports")
    regular_name = "MilanaReportRegular"
    bold_name = "MilanaReportBold"
    if regular_name not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont(regular_name, str(regular_path)))
    if bold_name not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont(bold_name, str(bold_path)))
    return regular_name, bold_name


def build_material_inventory_pdf(
    rows: list[dict],
    generated_label: str,
    lang: ReportLanguage,
    *,
    scope_label: str | None = None,
) -> bytes:
    text = REPORT_TEXT[lang]
    regular_font, bold_font = _register_report_fonts()
    output = BytesIO()
    page_size = landscape(A4)
    document = SimpleDocTemplate(
        output,
        pagesize=page_size,
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=13 * mm,
        bottomMargin=14 * mm,
        title=text["title"],
        author="Milana ERP",
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "InventoryReportTitle",
        parent=styles["Title"],
        fontName=bold_font,
        fontSize=16,
        leading=20,
        alignment=TA_LEFT,
        textColor=colors.HexColor("#1F1C17"),
        spaceAfter=4,
    )
    meta_style = ParagraphStyle(
        "InventoryReportMeta",
        parent=styles["Normal"],
        fontName=regular_font,
        fontSize=8.5,
        leading=11,
        textColor=colors.HexColor("#625B4B"),
        spaceAfter=2,
    )
    cell_style = ParagraphStyle(
        "InventoryReportCell",
        parent=styles["Normal"],
        fontName=regular_font,
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#1F1C17"),
    )
    header_style = ParagraphStyle(
        "InventoryReportHeader",
        parent=cell_style,
        fontName=bold_font,
        textColor=colors.white,
    )
    total_style = ParagraphStyle(
        "InventoryReportTotal",
        parent=cell_style,
        fontName=bold_font,
        alignment=TA_RIGHT,
    )

    elements = [
        Paragraph(escape(text["title"]), title_style),
        Paragraph(escape(text["subtitle"]), meta_style),
        Paragraph(
            escape(
                " | ".join(
                    value
                    for value in (f'{text["generated"]}: {generated_label}', scope_label)
                    if value
                )
            ),
            meta_style,
        ),
        Spacer(1, 5 * mm),
    ]
    table_data = [[
        Paragraph(escape(text["number"]), header_style),
        Paragraph(escape(text["material"]), header_style),
        Paragraph(escape(text["sku"]), header_style),
        Paragraph(escape(text["batches"]), header_style),
        Paragraph(escape(text["count"]), header_style),
        Paragraph(escape(text["kg"]), header_style),
    ]]
    for index, row in enumerate(rows, 1):
        table_data.append([
            index,
            Paragraph(escape(str(row["material_name"])), cell_style),
            Paragraph(escape(str(row["sku"])), cell_style),
            f'{int(row["batch_count"]):,}',
            f'{int(row["piece_count"]):,}',
            f'{float(row["total_kg"]):,.2f}',
        ])
    table_data.append([
        "",
        Paragraph(escape(text["grand_total"]), total_style),
        "",
        f'{sum(int(row["batch_count"]) for row in rows):,}',
        f'{sum(int(row["piece_count"]) for row in rows):,}',
        f'{sum(float(row["total_kg"]) for row in rows):,.2f}',
    ])
    last_row = len(table_data) - 1
    table = Table(
        table_data,
        colWidths=[12 * mm, 82 * mm, 48 * mm, 27 * mm, 34 * mm, 32 * mm],
        repeatRows=1,
        hAlign="LEFT",
    )
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F1C17")),
        ("FONTNAME", (0, 0), (-1, 0), bold_font),
        ("FONTNAME", (0, 1), (-1, -1), regular_font),
        ("FONTNAME", (0, last_row), (-1, last_row), bold_font),
        ("BACKGROUND", (0, last_row), (-1, last_row), colors.HexColor("#EEEAE0")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (0, -1), "RIGHT"),
        ("ALIGN", (3, 1), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, 0), 6),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ("TOPPADDING", (0, 1), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 4),
        ("LINEBELOW", (0, 0), (-1, -2), 0.35, colors.HexColor("#D8D1C0")),
        ("LINEABOVE", (0, last_row), (-1, last_row), 1, colors.HexColor("#1F1C17")),
    ]))
    elements.append(table)

    def draw_page_number(canvas, doc):
        canvas.saveState()
        canvas.setFont(regular_font, 8)
        canvas.setFillColor(colors.HexColor("#625B4B"))
        canvas.drawRightString(
            page_size[0] - document.rightMargin,
            7 * mm,
            f'{text["page"]} {doc.page}',
        )
        canvas.restoreState()

    document.build(elements, onFirstPage=draw_page_number, onLaterPages=draw_page_number)
    return output.getvalue()
