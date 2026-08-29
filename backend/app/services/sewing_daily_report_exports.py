from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Literal
from xml.sax.saxutils import escape
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.schemas.sewing_daily_report import SewingDailyReportListOut, SewingDailyReportOut


ReportLanguage = Literal["en", "ru", "uz"]

REPORT_TEXT = {
    "en": {
        "title": "Daily Sewing Report",
        "period": "Report period",
        "generated": "Generated",
        "summary": "Line summary",
        "entries": "Saved entries",
        "number": "No.",
        "date": "Report date",
        "saved_at": "Saved at",
        "line": "Sewing line",
        "section": "Section",
        "order": "Order",
        "model_no": "Model No.",
        "variant_no": "Variant No.",
        "kroy_no": "Kroy No.",
        "sewn": "Sewn qty",
        "defective": "Defect qty",
        "reason": "Defect reason",
        "notes": "Notes",
        "report_count": "Entries",
        "grand_total": "Grand total",
        "page": "Page",
        "none": "-",
    },
    "ru": {
        "title": "Ежедневный отчёт по швейному производству",
        "period": "Период отчёта",
        "generated": "Сформировано",
        "summary": "Итоги по линиям",
        "entries": "Сохранённые записи",
        "number": "№",
        "date": "Дата отчёта",
        "saved_at": "Время сохранения",
        "line": "Швейная линия",
        "section": "Секция",
        "order": "Заказ",
        "model_no": "Модель №",
        "variant_no": "Вариант №",
        "kroy_no": "Крой №",
        "sewn": "Сшито",
        "defective": "Брак",
        "reason": "Причина брака",
        "notes": "Примечания",
        "report_count": "Записи",
        "grand_total": "Общий итог",
        "page": "Страница",
        "none": "-",
    },
    "uz": {
        "title": "Kunlik tikuv hisoboti",
        "period": "Hisobot davri",
        "generated": "Yaratildi",
        "summary": "Liniyalar bo'yicha hisobot",
        "entries": "Saqlangan yozuvlar",
        "number": "№",
        "date": "Hisobot sanasi",
        "saved_at": "Saqlangan vaqt",
        "line": "Tikuv liniyasi",
        "section": "Qator",
        "order": "Buyurtma",
        "model_no": "Model raqami",
        "variant_no": "Variant raqami",
        "kroy_no": "Kroy No",
        "sewn": "Tikilgan miqdor",
        "defective": "Brak miqdori",
        "reason": "Brak sababi",
        "notes": "Izohlar",
        "report_count": "Yozuvlar",
        "grand_total": "Umumiy jami",
        "page": "Sahifa",
        "none": "-",
    },
}

DEFECT_REASON_TEXT = {
    "en": {
        "hole_present": "Hole present",
        "paint_stain_present": "Paint stain present",
        "cut_or_damaged": "Cut or damaged",
        "shorter_than_specified": "Shorter than specified",
    },
    "ru": {
        "hole_present": "Имеется отверстие",
        "paint_stain_present": "Имеется пятно краски",
        "cut_or_damaged": "Порезано или повреждено",
        "shorter_than_specified": "Короче заданного размера",
    },
    "uz": {
        "hole_present": "Teshik mavjud",
        "paint_stain_present": "Bo'yoq dog'i mavjud",
        "cut_or_damaged": "Kesilgan yoki shikastlangan",
        "shorter_than_specified": "Belgilangan o'lchamdan kalta",
    },
}

TASHKENT = ZoneInfo("Asia/Tashkent")


def _local_datetime(value: datetime) -> datetime:
    if value.tzinfo is None:
        value = value.replace(tzinfo=timezone.utc)
    return value.astimezone(TASHKENT).replace(tzinfo=None)


def _reason(value: str | None, lang: ReportLanguage) -> str:
    if not value:
        return REPORT_TEXT[lang]["none"]
    return DEFECT_REASON_TEXT[lang].get(value, value)


def _value(value: object, lang: ReportLanguage) -> str:
    text = str(value or "").strip()
    return text or REPORT_TEXT[lang]["none"]


def _period_label(report: SewingDailyReportListOut) -> str:
    if report.from_date == report.to_date:
        return report.from_date.isoformat()
    return f"{report.from_date.isoformat()} - {report.to_date.isoformat()}"


def _entry_values(row: SewingDailyReportOut, index: int, lang: ReportLanguage) -> list[object]:
    return [
        index,
        row.report_date,
        _local_datetime(row.created_at),
        f"{row.line_name} ({row.line_code})",
        row.section_name or row.section_no or REPORT_TEXT[lang]["none"],
        row.order_no or row.production_no or row.sales_order_no or REPORT_TEXT[lang]["none"],
        row.model_no or row.model_code or REPORT_TEXT[lang]["none"],
        row.variant_no or REPORT_TEXT[lang]["none"],
        row.kroy_no or REPORT_TEXT[lang]["none"],
        int(row.sewn_qty or 0),
        int(row.defective_qty or 0),
        _reason(row.defect_reason, lang),
        row.notes or REPORT_TEXT[lang]["none"],
    ]


def build_sewing_daily_report_xlsx(
    report: SewingDailyReportListOut,
    generated_label: str,
    lang: ReportLanguage,
) -> bytes:
    text = REPORT_TEXT[lang]
    workbook = Workbook()
    entries_sheet = workbook.active
    entries_sheet.title = "Entries"
    workbook.properties.title = text["title"]
    workbook.properties.subject = f'{text["period"]}: {_period_label(report)}'
    workbook.properties.creator = "Milana ERP"

    dark_fill = PatternFill("solid", fgColor="1F1C17")
    white_font = Font(color="FFFFFF", bold=True)
    title_font = Font(size=16, bold=True, color="1F1C17")
    muted_font = Font(size=9, color="625B4B")
    thin_bottom = Border(bottom=Side(style="thin", color="D8D1C0"))

    entries_sheet.merge_cells("A1:E1")
    entries_sheet["A1"] = text["entries"]
    entries_sheet["A1"].font = title_font
    entries_sheet.merge_cells("A2:E2")
    entries_sheet["A2"] = f'{text["period"]}: {_period_label(report)}'
    entries_sheet["A2"].font = muted_font
    entries_sheet.merge_cells("A3:E3")
    entries_sheet["A3"] = f'{text["generated"]}: {generated_label}'
    entries_sheet["A3"].font = muted_font

    entry_headers = [
        text["number"],
        text["line"],
        text["model_no"],
        text["kroy_no"],
        text["sewn"],
    ]
    for column, value in enumerate(entry_headers, 1):
        cell = entries_sheet.cell(5, column, value)
        cell.fill = dark_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_index, report_row in enumerate(report.rows, 6):
        values = [
            row_index - 5,
            f"{report_row.line_name} ({report_row.line_code})",
            report_row.model_no or report_row.model_code or text["none"],
            report_row.kroy_no or text["none"],
            int(report_row.sewn_qty or 0),
        ]
        for column, value in enumerate(values, 1):
            cell = entries_sheet.cell(row_index, column, value)
            cell.border = thin_bottom
            cell.alignment = Alignment(
                horizontal="right" if column in {1, 5} else "left",
                vertical="top",
                wrap_text=column == 2,
            )

    widths = {
        "A": 8,
        "B": 32,
        "C": 20,
        "D": 20,
        "E": 18,
    }
    for column, width in widths.items():
        entries_sheet.column_dimensions[column].width = width
    entries_sheet.freeze_panes = "A6"
    entries_sheet.auto_filter.ref = f"A5:E{max(5, 5 + len(report.rows))}"
    entries_sheet.sheet_view.showGridLines = False
    entries_sheet.page_setup.orientation = "landscape"
    entries_sheet.page_setup.fitToWidth = 1
    entries_sheet.page_margins.left = 0.25
    entries_sheet.page_margins.right = 0.25
    entries_sheet.page_margins.top = 0.4
    entries_sheet.page_margins.bottom = 0.4
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _register_report_fonts() -> tuple[str, str]:
    regular_candidates = [
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
        Path("C:/Windows/Fonts/arial.ttf"),
    ]
    bold_candidates = [
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
        Path("C:/Windows/Fonts/arialbd.ttf"),
    ]
    regular_path = next((path for path in regular_candidates if path.exists()), None)
    bold_path = next((path for path in bold_candidates if path.exists()), None)
    if not regular_path or not bold_path:
        raise RuntimeError("A Unicode TrueType font is required for daily sewing PDF reports")
    regular_name = "MilanaSewingReportRegular"
    bold_name = "MilanaSewingReportBold"
    if regular_name not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont(regular_name, str(regular_path)))
    if bold_name not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont(bold_name, str(bold_path)))
    return regular_name, bold_name


def build_sewing_daily_report_pdf(
    report: SewingDailyReportListOut,
    generated_label: str,
    lang: ReportLanguage,
) -> bytes:
    text = REPORT_TEXT[lang]
    regular_font, bold_font = _register_report_fonts()
    output = BytesIO()
    page_size = landscape(A4)
    document = SimpleDocTemplate(
        output,
        pagesize=page_size,
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=13 * mm,
        title=text["title"],
        author="Milana ERP",
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "SewingReportTitle",
        parent=styles["Title"],
        fontName=bold_font,
        fontSize=16,
        leading=20,
        alignment=TA_LEFT,
        textColor=colors.HexColor("#1F1C17"),
        spaceAfter=4,
    )
    section_style = ParagraphStyle(
        "SewingReportSection",
        parent=styles["Heading2"],
        fontName=bold_font,
        fontSize=10,
        leading=13,
        textColor=colors.HexColor("#1F1C17"),
        spaceBefore=6,
        spaceAfter=4,
    )
    meta_style = ParagraphStyle(
        "SewingReportMeta",
        parent=styles["Normal"],
        fontName=regular_font,
        fontSize=8.5,
        leading=11,
        textColor=colors.HexColor("#625B4B"),
        spaceAfter=2,
    )
    cell_style = ParagraphStyle(
        "SewingReportCell",
        parent=styles["Normal"],
        fontName=regular_font,
        fontSize=6.7,
        leading=8,
        textColor=colors.HexColor("#1F1C17"),
    )
    header_style = ParagraphStyle(
        "SewingReportHeader",
        parent=cell_style,
        fontName=bold_font,
        textColor=colors.white,
    )
    total_style = ParagraphStyle(
        "SewingReportTotal",
        parent=cell_style,
        fontName=bold_font,
        alignment=TA_RIGHT,
    )

    elements = [
        Paragraph(escape(text["title"]), title_style),
        Paragraph(escape(f'{text["period"]}: {_period_label(report)}'), meta_style),
        Paragraph(escape(f'{text["generated"]}: {generated_label}'), meta_style),
        Spacer(1, 3 * mm),
        Paragraph(escape(text["summary"]), section_style),
    ]

    summary_data = [[
        Paragraph(escape(text["line"]), header_style),
        Paragraph(escape(text["sewn"]), header_style),
        Paragraph(escape(text["defective"]), header_style),
        Paragraph(escape(text["report_count"]), header_style),
        Paragraph(escape(text["model_no"]), header_style),
        Paragraph(escape(text["kroy_no"]), header_style),
    ]]
    for line in report.summary:
        model_labels = sorted({
            " / ".join(filter(None, [model.model_no or model.model_code, model.variant_no]))
            for model in line.models
            if model.model_no or model.model_code or model.variant_no
        })
        summary_data.append([
            Paragraph(escape(f"{line.line_name} ({line.line_code})"), cell_style),
            f"{int(line.total_sewn_qty or 0):,}",
            f"{int(line.total_defective_qty or 0):,}",
            f"{int(line.report_count or 0):,}",
            Paragraph(escape(", ".join(model_labels) or text["none"]), cell_style),
            Paragraph(escape(", ".join(line.kroy_nos) or text["none"]), cell_style),
        ])
    summary_data.append([
        Paragraph(escape(text["grand_total"]), total_style),
        f"{int(report.total_sewn_qty or 0):,}",
        f"{int(report.total_defective_qty or 0):,}",
        f"{len(report.rows):,}",
        "",
        "",
    ])
    summary_table = Table(
        summary_data,
        colWidths=[46 * mm, 22 * mm, 22 * mm, 19 * mm, 72 * mm, 50 * mm],
        repeatRows=1,
        hAlign="LEFT",
    )
    summary_table.setStyle(_table_style(regular_font, bold_font, len(summary_data) - 1))
    elements.extend([summary_table, Spacer(1, 4 * mm), Paragraph(escape(text["entries"]), section_style)])

    detail_data = [[
        Paragraph(escape(text["number"]), header_style),
        Paragraph(escape(text["date"]), header_style),
        Paragraph(escape(text["saved_at"]), header_style),
        Paragraph(escape(text["line"]), header_style),
        Paragraph(escape(text["section"]), header_style),
        Paragraph(escape(text["order"]), header_style),
        Paragraph(escape(text["model_no"]), header_style),
        Paragraph(escape(text["variant_no"]), header_style),
        Paragraph(escape(text["kroy_no"]), header_style),
        Paragraph(escape(text["sewn"]), header_style),
        Paragraph(escape(text["defective"]), header_style),
        Paragraph(escape(text["reason"]), header_style),
        Paragraph(escape(text["notes"]), header_style),
    ]]
    for index, row in enumerate(report.rows, 1):
        raw = _entry_values(row, index, lang)
        raw[1] = row.report_date.isoformat()
        raw[2] = _local_datetime(row.created_at).strftime("%Y-%m-%d %H:%M")
        detail_data.append([
            value if column in {0, 9, 10} else Paragraph(escape(str(value)), cell_style)
            for column, value in enumerate(raw)
        ])
    detail_table = Table(
        detail_data,
        colWidths=[
            10 * mm,
            17 * mm,
            25 * mm,
            31 * mm,
            18 * mm,
            21 * mm,
            21 * mm,
            17 * mm,
            18 * mm,
            13 * mm,
            13 * mm,
            28 * mm,
            36 * mm,
        ],
        repeatRows=1,
        hAlign="LEFT",
    )
    detail_table.setStyle(_table_style(regular_font, bold_font, None))
    elements.append(detail_table)

    def draw_page_number(canvas, doc):
        canvas.saveState()
        canvas.setFont(regular_font, 8)
        canvas.setFillColor(colors.HexColor("#625B4B"))
        canvas.drawRightString(
            page_size[0] - document.rightMargin,
            6 * mm,
            f'{text["page"]} {doc.page}',
        )
        canvas.restoreState()

    document.build(elements, onFirstPage=draw_page_number, onLaterPages=draw_page_number)
    return output.getvalue()


def _table_style(regular_font: str, bold_font: str, total_row: int | None) -> TableStyle:
    rules: list[tuple] = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F1C17")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), bold_font),
        ("FONTNAME", (0, 1), (-1, -1), regular_font),
        ("ALIGN", (0, 0), (0, -1), "RIGHT"),
        ("ALIGN", (1, 1), (3, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTSIZE", (0, 0), (-1, -1), 6.7),
        ("TOPPADDING", (0, 0), (-1, 0), 5),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 5),
        ("TOPPADDING", (0, 1), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 3),
        ("LINEBELOW", (0, 0), (-1, -1), 0.3, colors.HexColor("#D8D1C0")),
    ]
    if total_row is not None:
        rules.extend([
            ("BACKGROUND", (0, total_row), (-1, total_row), colors.HexColor("#EEEAE0")),
            ("FONTNAME", (0, total_row), (-1, total_row), bold_font),
            ("LINEABOVE", (0, total_row), (-1, total_row), 1, colors.HexColor("#1F1C17")),
        ])
    return TableStyle(rules)
