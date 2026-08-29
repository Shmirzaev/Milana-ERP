from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any, Literal
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ReportLanguage = Literal["en", "ru", "uz"]
TASHKENT = ZoneInfo("Asia/Tashkent")

REPORT_TEXT: dict[ReportLanguage, dict[str, str]] = {
    "en": {
        "title": "Daily attendance report",
        "date": "Date",
        "generated": "Generated",
        "summary": "Profiles: {total}  Complete: {complete}  One scan: {single}  Absent: {absent}",
        "number": "No.",
        "employee_id": "Employee ID",
        "employee": "Employee",
        "status": "Status",
        "arrival": "Arrival (first scan)",
        "departure": "Departure (last scan)",
        "worked": "Time between scans",
        "complete": "Arrival and departure",
        "single_scan": "One scan only",
        "absent": "No scans",
    },
    "ru": {
        "title": "Ежедневный отчёт о посещаемости",
        "date": "Дата",
        "generated": "Сформировано",
        "summary": "Профилей: {total}  Полный день: {complete}  Один проход: {single}  Нет проходов: {absent}",
        "number": "№",
        "employee_id": "ID сотрудника",
        "employee": "Сотрудник",
        "status": "Статус",
        "arrival": "Приход (первый проход)",
        "departure": "Уход (последний проход)",
        "worked": "Время между проходами",
        "complete": "Приход и уход",
        "single_scan": "Только один проход",
        "absent": "Нет проходов",
    },
    "uz": {
        "title": "Kunlik davomat hisoboti",
        "date": "Sana",
        "generated": "Yaratilgan vaqt",
        "summary": "Profillar: {total}  To‘liq: {complete}  Bitta o‘tish: {single}  Kelmagan: {absent}",
        "number": "№",
        "employee_id": "Xodim IDsi",
        "employee": "Xodim",
        "status": "Holat",
        "arrival": "Kelish (birinchi o‘tish)",
        "departure": "Ketish (oxirgi o‘tish)",
        "worked": "O‘tishlar orasidagi vaqt",
        "complete": "Kelish va ketish",
        "single_scan": "Faqat bitta o‘tish",
        "absent": "O‘tish yo‘q",
    },
}


def _local_time(value: datetime | None) -> str:
    if value is None:
        return ""
    return value.astimezone(TASHKENT).strftime("%H:%M")


def _duration(minutes: int | None) -> str:
    if minutes is None:
        return ""
    hours, remainder = divmod(max(minutes, 0), 60)
    return f"{hours:02d}:{remainder:02d}"


def build_daily_attendance_xlsx(
    *,
    day: date,
    rows: list[dict[str, Any]],
    generated_at: datetime,
    lang: ReportLanguage,
) -> bytes:
    text = REPORT_TEXT[lang]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Attendance"
    workbook.properties.title = text["title"]

    dark_fill = PatternFill("solid", fgColor="1F1C17")
    light_fill = PatternFill("solid", fgColor="EEEAE0")
    white_font = Font(color="FFFFFF", bold=True)
    title_font = Font(size=16, bold=True, color="1F1C17")
    thin_border = Border(bottom=Side(style="thin", color="D8D1C0"))

    sheet.merge_cells("A1:G1")
    sheet["A1"] = text["title"]
    sheet["A1"].font = title_font
    sheet.merge_cells("A2:G2")
    sheet["A2"] = f'{text["date"]}: {day.isoformat()}'
    sheet["A2"].font = Font(size=10, color="625B4B")

    counts = {"complete": 0, "single_scan": 0, "absent": 0}
    for row in rows:
        counts[str(row["attendance_status"])] += 1
    sheet.merge_cells("A3:G3")
    sheet["A3"] = text["summary"].format(
        total=len(rows),
        complete=counts["complete"],
        single=counts["single_scan"],
        absent=counts["absent"],
    )
    sheet["A3"].font = Font(size=10, color="625B4B")
    sheet.merge_cells("A4:G4")
    sheet["A4"] = f'{text["generated"]}: {generated_at.astimezone(TASHKENT).strftime("%Y-%m-%d %H:%M")}'
    sheet["A4"].font = Font(size=9, color="7A725F")

    header_row = 6
    headers = [
        text["number"],
        text["employee_id"],
        text["employee"],
        text["status"],
        text["arrival"],
        text["departure"],
        text["worked"],
    ]
    for column, value in enumerate(headers, 1):
        cell = sheet.cell(header_row, column, value)
        cell.fill = dark_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for index, row in enumerate(rows, 1):
        excel_row = header_row + index
        status = str(row["attendance_status"])
        values = [
            index,
            row["external_person_id"],
            row["full_name"],
            text[status],
            _local_time(row.get("arrival_at")),
            _local_time(row.get("departure_at")),
            _duration(row.get("worked_minutes")),
        ]
        for column, value in enumerate(values, 1):
            cell = sheet.cell(excel_row, column, value)
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal="center" if column in {1, 2, 4, 5, 6, 7} else "left",
                vertical="center",
            )
        if status == "absent":
            for column in range(1, 8):
                sheet.cell(excel_row, column).fill = light_fill

    widths = {"A": 7, "B": 18, "C": 42, "D": 24, "E": 22, "F": 22, "G": 24}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.row_dimensions[header_row].height = 32
    sheet.freeze_panes = "A7"
    sheet.auto_filter.ref = f"A{header_row}:G{max(header_row, header_row + len(rows))}"
    sheet.sheet_view.showGridLines = False
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.4
    sheet.page_margins.bottom = 0.4

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
