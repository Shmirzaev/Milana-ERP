from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Literal
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ReportLanguage = Literal["en", "ru", "uz"]
REPORT_TIMEZONE = ZoneInfo("Asia/Tashkent")

REPORT_TEXT = {
    "en": {
        "title": "Sewing production report",
        "period": "Reporting period",
        "generated": "Generated",
        "number": "No.",
        "date": "Date",
        "employee": "Employee",
        "employee_no": "Employee no.",
        "barcode": "QR number",
        "line": "Sewing line",
        "cutting": "Cutting / batch",
        "model": "Sewing model",
        "product": "Product",
        "operation": "Operation name",
        "size": "Size",
        "quantity": "Completed work",
        "rate": "Average rate",
        "amount": "Amount",
        "currency": "Currency",
        "totals": "Totals",
        "qr_count": "Scanned QR codes",
        "total_quantity": "Total completed pieces",
        "total_rate": "Total rate",
        "total_amount": "Total amount",
    },
    "ru": {
        "title": "Отчёт по швейному производству",
        "period": "Отчётный период",
        "generated": "Сформировано",
        "number": "№",
        "date": "Дата",
        "employee": "Сотрудник",
        "employee_no": "Табельный номер",
        "barcode": "Номер QR",
        "line": "Швейный поток",
        "cutting": "Крой / партия",
        "model": "Швейная модель",
        "product": "Продукт",
        "operation": "Наименование операции",
        "size": "Размер",
        "quantity": "Выполнено работ",
        "rate": "Средняя цена",
        "amount": "Сумма",
        "currency": "Валюта",
        "totals": "Итого",
        "qr_count": "Отсканировано QR-кодов",
        "total_quantity": "Всего выполнено работ",
        "total_rate": "Общая сумма расценок",
        "total_amount": "Общая сумма",
    },
    "uz": {
        "title": "Tikuv ishlab chiqarish hisoboti",
        "period": "Hisobot davri",
        "generated": "Yaratildi",
        "number": "№",
        "date": "Sana",
        "employee": "Ishchi",
        "employee_no": "Xodim raqami",
        "barcode": "QR raqami",
        "line": "Tikuv potogi",
        "cutting": "Bichuv / partiya",
        "model": "Tikuv modeli",
        "product": "Mahsulot",
        "operation": "Operatsiya nomi",
        "size": "O‘lcham",
        "quantity": "Bajarilgan ishlar soni",
        "rate": "O‘rtacha narx",
        "amount": "Summa",
        "currency": "Valyuta",
        "totals": "Jami",
        "qr_count": "Skanerlangan QR kodlar",
        "total_quantity": "Jami bajarilgan ishlar",
        "total_rate": "Jami narx",
        "total_amount": "Umumiy summa",
    },
}


def build_sewing_production_report_xlsx(
    rows: list[dict],
    *,
    date_from: datetime | None,
    date_to: datetime | None,
    generated_label: str,
    lang: ReportLanguage,
    currency: str,
) -> bytes:
    text = REPORT_TEXT[lang]

    def safe_cell_value(value):
        if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
            return f"'{value}"
        return value

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sewing report"
    workbook.properties.title = text["title"]

    dark_fill = PatternFill("solid", fgColor="1F1C17")
    total_fill = PatternFill("solid", fgColor="EEEAE0")
    white_font = Font(color="FFFFFF", bold=True)
    title_font = Font(size=16, bold=True, color="1F1C17")
    thin_side = Side(style="thin", color="D8D1C0")
    row_border = Border(bottom=thin_side)

    sheet.merge_cells("A1:O1")
    sheet["A1"] = text["title"]
    sheet["A1"].font = title_font
    sheet.merge_cells("A2:O2")
    def local_datetime(value: datetime | None) -> datetime | None:
        if value is None:
            return None
        if value.tzinfo is None:
            return value
        return value.astimezone(REPORT_TIMEZONE).replace(tzinfo=None)

    local_date_from = local_datetime(date_from)
    local_date_to = local_datetime(date_to)
    period_from = local_date_from.strftime("%Y-%m-%d %H:%M:%S") if local_date_from else "—"
    period_to = local_date_to.strftime("%Y-%m-%d %H:%M:%S") if local_date_to else "—"
    sheet["A2"] = f'{text["period"]}: {period_from} — {period_to}'
    sheet["A2"].font = Font(size=10, color="625B4B")
    sheet.merge_cells("A3:O3")
    sheet["A3"] = f'{text["generated"]}: {generated_label}'
    sheet["A3"].font = Font(size=9, color="7A725F")

    header_row = 5
    headers = [
        text["number"],
        text["date"],
        text["employee"],
        text["employee_no"],
        text["barcode"],
        text["line"],
        text["cutting"],
        text["model"],
        text["product"],
        text["operation"],
        text["size"],
        text["quantity"],
        text["rate"],
        text["amount"],
        text["currency"],
    ]
    for column, value in enumerate(headers, 1):
        cell = sheet.cell(header_row, column, value)
        cell.fill = dark_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for index, row in enumerate(rows, 1):
        excel_row = header_row + index
        scanned_at = row.get("scanned_at")
        if isinstance(scanned_at, datetime):
            scanned_at = local_datetime(scanned_at)
        line = " — ".join(
            str(value).strip()
            for value in (row.get("sewing_line_code"), row.get("sewing_line_name"))
            if value and str(value).strip()
        )
        values = [
            index,
            scanned_at,
            row.get("employee_name") or "",
            row.get("employee_no") or row.get("employee_id") or "",
            row.get("barcode") or "",
            line,
            row.get("cutting_reference") or "",
            row.get("model_code") or "",
            row.get("product_name") or "",
            row.get("operation_name") or row.get("operation_code") or "",
            row.get("size") or "",
            float(row.get("quantity") or 0),
            float(row.get("rate_per_piece") or 0),
            float(row.get("total_amount") or 0),
            row.get("currency") or currency,
        ]
        for column, value in enumerate(values, 1):
            cell = sheet.cell(excel_row, column, safe_cell_value(value))
            cell.border = row_border
            cell.alignment = Alignment(
                horizontal="right" if column in {1, 12, 13, 14} else "left",
                vertical="top",
                wrap_text=column in {3, 6, 7, 8, 9, 10},
            )
        sheet.cell(excel_row, 2).number_format = "yyyy-mm-dd hh:mm:ss"
        for column in (12, 13, 14):
            sheet.cell(excel_row, column).number_format = '#,##0.00'

    totals_start = header_row + len(rows) + 2
    total_quantity = sum(float(row.get("quantity") or 0) for row in rows)
    total_rate = sum(float(row.get("rate_per_piece") or 0) for row in rows)
    total_amount = sum(float(row.get("total_amount") or 0) for row in rows)
    summary_rows = [
        (text["qr_count"], len(rows), "0"),
        (text["total_quantity"], total_quantity, '#,##0.00'),
        (text["total_rate"], total_rate, '#,##0.00'),
        (text["total_amount"], total_amount, '#,##0.00'),
    ]
    sheet.merge_cells(start_row=totals_start, start_column=1, end_row=totals_start, end_column=3)
    sheet.cell(totals_start, 1, text["totals"])
    for column in range(1, 4):
        cell = sheet.cell(totals_start, column)
        cell.fill = dark_fill
        cell.font = white_font

    for row_offset, (label, value, number_format) in enumerate(summary_rows, 1):
        summary_row = totals_start + row_offset
        sheet.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=2)
        sheet.cell(summary_row, 1, label)
        sheet.cell(summary_row, 3, value)
        sheet.cell(summary_row, 3).number_format = number_format
        for column in range(1, 4):
            cell = sheet.cell(summary_row, column)
            cell.fill = total_fill
            cell.font = Font(bold=True, color="1F1C17")
            cell.border = row_border
        if row_offset in {3, 4}:
            sheet.cell(summary_row, 4, currency)
            sheet.cell(summary_row, 4).fill = total_fill
            sheet.cell(summary_row, 4).font = Font(bold=True, color="1F1C17")

    widths = {
        "A": 7,
        "B": 20,
        "C": 26,
        "D": 16,
        "E": 18,
        "F": 24,
        "G": 22,
        "H": 20,
        "I": 28,
        "J": 34,
        "K": 12,
        "L": 18,
        "M": 16,
        "N": 18,
        "O": 12,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.row_dimensions[header_row].height = 34
    sheet.freeze_panes = "A6"
    sheet.auto_filter.ref = f"A{header_row}:O{max(header_row, header_row + len(rows))}"
    sheet.sheet_view.showGridLines = False
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.4
    sheet.page_margins.bottom = 0.4

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
