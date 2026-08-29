from io import BytesIO
from types import SimpleNamespace
from uuid import uuid4

from openpyxl import load_workbook
from sqlalchemy.dialects import postgresql

from app.api.routes.cutting_passports import _compute, _size_count_from_range
from app.api.routes.inventory import (
    _delete_stock_batch_receipt_movements,
    _locked_active_batch_reservations_statement,
    _locked_stock_batch_statement,
)


def test_cutting_passport_bitta_ish_gr_uses_size_count():
    passport = SimpleNamespace(
        pieces=600,
        beka_per_piece_kg=0.005,
        other_beka_per_piece_kg=0.001,
        ribana_per_piece_kg=0,
        scrap_kg=2,
        layer_weight_kg=0,
        total_layers=50,
        fabric_width_m=1.8,
        lay_length_m=3.37,
        gramage=0.191,
        planned_kg=0,
        size_range="46-56",
    )

    result = _compute(passport)

    assert _size_count_from_range("46-56") == 6
    assert _size_count_from_range("44-52") == 5
    assert result["pieces_per_layer"] == 12
    assert result["size_count"] == 6
    bitta_ish_gr = round((1.8 * 3.37 * 0.191) / 6 + 0.005 + 0.001, 6)
    assert result["per_piece_weight_kg"] == bitta_ish_gr
    assert result["theoretical_kg"] == round(bitta_ish_gr * 600 + 2, 6)


def test_admin_can_create_update_and_delete_inventory_item(client, auth_headers):
    create = client.post(
        "/api/inventory/items",
        json={
            "sku": "FAB-TEST-DELETE",
            "name": "Test fabric before edit",
            "category": "fabric",
            "unit": "kg",
            "default_cost": 1.25,
            "reorder_level": 10,
            "track_batch": True,
            "is_active": True,
            "composition": [
                {"name": "Cotton", "percentage": 95},
                {"name": "Elastane", "percentage": 5},
            ],
        },
        headers=auth_headers,
    )
    assert create.status_code == 201, create.text
    created_item = create.json()
    item_id = created_item["id"]
    assert created_item["composition"] == [
        {"name": "Cotton", "percentage": 95.0},
        {"name": "Elastane", "percentage": 5.0},
    ]

    update = client.patch(
        f"/api/inventory/items/{item_id}",
        json={
            "sku": "FAB-TEST-DELETE-EDIT",
            "name": "Test fabric after edit",
            "category": "fabric",
            "unit": "m",
            "default_cost": 2.5,
            "reorder_level": 20,
            "track_batch": False,
            "is_active": True,
            "composition": [{"name": "Viscose", "percentage": 100}],
        },
        headers=auth_headers,
    )
    assert update.status_code == 200, update.text
    assert update.json()["sku"] == "FAB-TEST-DELETE-EDIT"
    assert update.json()["unit"] == "m"
    assert update.json()["composition"] == [{"name": "Viscose", "percentage": 100.0}]

    delete = client.delete(f"/api/inventory/items/{item_id}", headers=auth_headers)
    assert delete.status_code == 204, delete.text

    items = client.get("/api/inventory/items?q=FAB-TEST-DELETE-EDIT", headers=auth_headers)
    assert items.status_code == 200, items.text
    assert all(row["id"] != item_id for row in items.json())


def test_inventory_item_rejects_composition_total_over_100(client, auth_headers):
    response = client.post(
        "/api/inventory/items",
        json={
            "sku": "FAB-COMPOSITION-OVER",
            "name": "Invalid composition fabric",
            "category": "fabric",
            "unit": "kg",
            "default_cost": 1,
            "reorder_level": 0,
            "track_batch": True,
            "is_active": True,
            "composition": [
                {"name": "Cotton", "percentage": 80},
                {"name": "Polyester", "percentage": 30},
            ],
        },
        headers=auth_headers,
    )

    assert response.status_code == 400, response.text
    assert response.json()["detail"] == "Composition total cannot exceed 100%"


def test_inventory_item_rejects_duplicate_active_material_name(client, auth_headers):
    suffix = uuid4().hex[:8].upper()
    create = client.post(
        "/api/inventory/items",
        json={
            "sku": f"FAB-DUP-NAME-A-{suffix}",
            "name": f"Suprem duplicate {suffix}",
            "category": "fabric",
            "unit": "kg",
            "default_cost": 1,
            "reorder_level": 0,
            "track_batch": True,
            "is_active": True,
            "composition": [{"name": "Cotton", "percentage": 100}],
        },
        headers=auth_headers,
    )
    assert create.status_code == 201, create.text

    duplicate = client.post(
        "/api/inventory/items",
        json={
            "sku": f"FAB-DUP-NAME-B-{suffix}",
            "name": f"  suprem duplicate {suffix}  ",
            "category": "fabric",
            "unit": "kg",
            "default_cost": 1,
            "reorder_level": 0,
            "track_batch": True,
            "is_active": True,
            "composition": [{"name": "Cotton", "percentage": 100}],
        },
        headers=auth_headers,
    )
    assert duplicate.status_code == 400, duplicate.text
    assert duplicate.json()["detail"] == "Material name already exists"


def test_inventory_item_update_rejects_duplicate_active_material_name(client, auth_headers):
    suffix = uuid4().hex[:8].upper()
    first = client.post(
        "/api/inventory/items",
        json={
            "sku": f"FAB-DUP-UPDATE-A-{suffix}",
            "name": f"Merge fabric {suffix}",
            "category": "fabric",
            "unit": "kg",
            "default_cost": 1,
            "reorder_level": 0,
            "track_batch": True,
            "is_active": True,
        },
        headers=auth_headers,
    )
    assert first.status_code == 201, first.text
    second = client.post(
        "/api/inventory/items",
        json={
            "sku": f"FAB-DUP-UPDATE-B-{suffix}",
            "name": f"Other fabric {suffix}",
            "category": "fabric",
            "unit": "kg",
            "default_cost": 1,
            "reorder_level": 0,
            "track_batch": True,
            "is_active": True,
        },
        headers=auth_headers,
    )
    assert second.status_code == 201, second.text

    duplicate_update = client.patch(
        f"/api/inventory/items/{second.json()['id']}",
        json={
            "sku": f"FAB-DUP-UPDATE-B-{suffix}",
            "name": f"merge fabric {suffix}",
            "category": "fabric",
            "unit": "kg",
            "default_cost": 1,
            "reorder_level": 0,
            "track_batch": True,
            "is_active": True,
        },
        headers=auth_headers,
    )
    assert duplicate_update.status_code == 400, duplicate_update.text
    assert duplicate_update.json()["detail"] == "Material name already exists"


def test_cannot_delete_item_linked_to_stock(client, auth_headers):
    create = client.post(
        "/api/inventory/items",
        json={
            "sku": "FAB-TEST-LINKED",
            "name": "Linked fabric",
            "category": "fabric",
            "unit": "kg",
            "default_cost": 1,
            "reorder_level": 0,
            "track_batch": True,
            "is_active": True,
        },
        headers=auth_headers,
    )
    assert create.status_code == 201, create.text
    item = create.json()

    warehouses = client.get("/api/inventory/warehouses", headers=auth_headers)
    assert warehouses.status_code == 200, warehouses.text
    warehouse_id = next(row["id"] for row in warehouses.json() if row["type"] == "fabric_storage")

    receive = client.post(
        "/api/inventory/receive",
        json={
            "item_id": item["id"],
            "batch_no": "LINKED-ITEM-001",
            "quantity": 1,
            "unit": "kg",
            "cost_per_unit": 1,
            "warehouse_id": warehouse_id,
            "qc_status": "passed",
        },
        headers=auth_headers,
    )
    assert receive.status_code == 201, receive.text

    delete = client.delete(f"/api/inventory/items/{item['id']}", headers=auth_headers)
    assert delete.status_code == 409, delete.text


def test_inventory_batches_can_filter_by_material_item_ids(client, auth_headers):
    fabric = client.post(
        "/api/inventory/items",
        json={
            "sku": "FAB-BATCH-FILTER",
            "name": "Batch filter fabric",
            "category": "fabric",
            "unit": "kg",
            "default_cost": 3,
            "reorder_level": 0,
            "track_batch": True,
            "is_active": True,
        },
        headers=auth_headers,
    )
    assert fabric.status_code == 201, fabric.text
    accessory = client.post(
        "/api/inventory/items",
        json={
            "sku": "ACC-BATCH-FILTER",
            "name": "Batch filter accessory",
            "category": "accessory",
            "unit": "pcs",
            "default_cost": 1,
            "reorder_level": 0,
            "track_batch": True,
            "is_active": True,
        },
        headers=auth_headers,
    )
    assert accessory.status_code == 201, accessory.text

    warehouses = client.get("/api/inventory/warehouses", headers=auth_headers)
    assert warehouses.status_code == 200, warehouses.text
    fabric_warehouse_id = next(row["id"] for row in warehouses.json() if row["type"] == "fabric_storage")
    accessory_warehouse_id = next(row["id"] for row in warehouses.json() if row["type"] == "accessory_storage")

    fabric_batch = client.post(
        "/api/inventory/receive",
        json={
            "item_id": fabric.json()["id"],
            "batch_no": "FAB-BATCH-FILTER-001",
            "quantity": 12,
            "unit": "kg",
            "cost_per_unit": 3,
            "warehouse_id": fabric_warehouse_id,
            "qc_status": "passed",
        },
        headers=auth_headers,
    )
    assert fabric_batch.status_code == 201, fabric_batch.text
    accessory_batch = client.post(
        "/api/inventory/receive",
        json={
            "item_id": accessory.json()["id"],
            "batch_no": "ACC-BATCH-FILTER-001",
            "quantity": 50,
            "unit": "pcs",
            "cost_per_unit": 1,
            "warehouse_id": accessory_warehouse_id,
            "qc_status": "passed",
        },
        headers=auth_headers,
    )
    assert accessory_batch.status_code == 201, accessory_batch.text

    by_item = client.get(
        f"/api/inventory/batches?item_ids={fabric.json()['id']}",
        headers=auth_headers,
    )
    assert by_item.status_code == 200, by_item.text
    by_item_batch_nos = {row["batch_no"] for row in by_item.json()}
    assert "FAB-BATCH-FILTER-001" in by_item_batch_nos
    assert "ACC-BATCH-FILTER-001" not in by_item_batch_nos
    fabric_row = next(row for row in by_item.json() if row["batch_no"] == "FAB-BATCH-FILTER-001")
    assert fabric_row["item_category"] == "fabric"

    material_group = client.get("/api/inventory/batches?group=materials&page_size=1000", headers=auth_headers)
    assert material_group.status_code == 200, material_group.text
    material_batch_nos = {row["batch_no"] for row in material_group.json()}
    assert "FAB-BATCH-FILTER-001" in material_batch_nos
    assert "ACC-BATCH-FILTER-001" not in material_batch_nos


def test_material_inventory_batch_details_include_receive_fields_and_search(client, auth_headers):
    suffix = uuid4().hex[:8].upper()
    item_response = client.post(
        "/api/inventory/items",
        json={
            "sku": f"FAB-RECV-VISIBLE-{suffix}",
            "name": "Receive visible fabric",
            "category": "fabric",
            "unit": "kg",
            "default_cost": 4.2,
            "reorder_level": 0,
            "track_batch": True,
            "is_active": True,
        },
        headers=auth_headers,
    )
    assert item_response.status_code == 201, item_response.text
    item = item_response.json()

    supplier_response = client.post(
        "/api/suppliers",
        json={"name": f"Visible Supplier {suffix}", "phone": "+998901234567"},
        headers=auth_headers,
    )
    assert supplier_response.status_code == 201, supplier_response.text
    supplier = supplier_response.json()

    warehouses_response = client.get("/api/inventory/warehouses", headers=auth_headers)
    assert warehouses_response.status_code == 200, warehouses_response.text
    warehouse = next(row for row in warehouses_response.json() if row["type"] == "fabric_storage")

    receive_payload = {
        "item_id": item["id"],
        "batch_no": f"BATCH-VISIBLE-{suffix}",
        "supplier_id": supplier["id"],
        "color": "Ocean green",
        "old_code": f"OLD-{suffix}",
        "color_code": f"CLR-{suffix}",
        "color_status": "Approved shade",
        "order_no": f"ORDER-VISIBLE-{suffix}",
        "quantity": 123.45,
        "gsm": 0.145678,
        "piece_count": 17,
        "processes": f"Wash, print, inspect {suffix}",
        "unit": "kg",
        "cost_per_unit": 6.75,
        "image_url": f"/storage/model-files/batch-visible-{suffix}.webp",
        "warehouse_id": warehouse["id"],
        "qc_status": "passed",
    }
    receive_response = client.post("/api/inventory/receive", json=receive_payload, headers=auth_headers)
    assert receive_response.status_code == 201, receive_response.text

    stock_response = client.get(
        f"/api/inventory/stock?group=materials&q=ORDER-VISIBLE-{suffix}",
        headers=auth_headers,
    )
    assert stock_response.status_code == 200, stock_response.text
    assert item["sku"] in {row["item_sku"] for row in stock_response.json()}

    batch_stock_response = client.get(
        f"/api/inventory/stock?group=materials&q=BATCH-VISIBLE-{suffix}",
        headers=auth_headers,
    )
    assert batch_stock_response.status_code == 200, batch_stock_response.text
    assert item["sku"] in {row["item_sku"] for row in batch_stock_response.json()}

    batches_response = client.get(
        f"/api/inventory/batches?group=materials&q=ORDER-VISIBLE-{suffix}&include_total=true",
        headers=auth_headers,
    )
    assert batches_response.status_code == 200, batches_response.text
    body = batches_response.json()
    batch = next(row for row in body["rows"] if row["batch_no"] == receive_payload["batch_no"])
    assert body["total"] >= 1
    assert batch["item_id"] == item["id"]
    assert batch["item_sku"] == item["sku"]
    assert batch["item_name"] == item["name"]
    assert batch["supplier_id"] == supplier["id"]
    assert batch["supplier_name"] == supplier["name"]
    assert batch["color"] == receive_payload["color"]
    assert batch["old_code"] == receive_payload["old_code"]
    assert batch["color_code"] == receive_payload["color_code"]
    assert batch["color_status"] == receive_payload["color_status"]
    assert batch["order_no"] == receive_payload["order_no"]
    assert round(float(batch["quantity"]), 2) == 123.45
    assert round(float(batch["gsm"]), 6) == 0.145678
    assert batch["piece_count"] == receive_payload["piece_count"]
    assert batch["processes"] == receive_payload["processes"]
    assert batch["unit"] == receive_payload["unit"]
    assert round(float(batch["cost_per_unit"]), 2) == 6.75
    assert batch["image_url"] == receive_payload["image_url"]
    assert batch["warehouse_id"] == warehouse["id"]
    assert batch["warehouse_name"] == warehouse["name"]
    assert batch["qc_status"] == receive_payload["qc_status"]
    assert round(float(batch["available_quantity"]), 2) == 123.45

    colors_response = client.get("/api/inventory/colors", headers=auth_headers)
    assert colors_response.status_code == 200, colors_response.text
    assert receive_payload["color"] in colors_response.json()


def test_material_inventory_search_ignores_depleted_batch_matches(client, auth_headers):
    suffix = uuid4().hex[:8].upper()
    stale_batch_no = f"DEPLETED-SEARCH-{suffix}"
    live_batch_no = f"LIVE-SEARCH-{suffix}"
    item_response = client.post(
        "/api/inventory/items",
        json={
            "sku": f"FAB-SEARCH-CONTROL-{suffix}",
            "name": f"Batch search control fabric {suffix}",
            "category": "fabric",
            "unit": "kg",
            "default_cost": 1,
            "reorder_level": 0,
            "track_batch": True,
            "is_active": True,
        },
        headers=auth_headers,
    )
    assert item_response.status_code == 201, item_response.text
    item = item_response.json()

    warehouses_response = client.get("/api/inventory/warehouses", headers=auth_headers)
    assert warehouses_response.status_code == 200, warehouses_response.text
    warehouse = next(row for row in warehouses_response.json() if row["type"] == "fabric_storage")

    stale_receive = client.post(
        "/api/inventory/receive",
        json={
            "item_id": item["id"],
            "batch_no": stale_batch_no,
            "quantity": 10,
            "unit": "kg",
            "cost_per_unit": 1,
            "warehouse_id": warehouse["id"],
            "qc_status": "passed",
        },
        headers=auth_headers,
    )
    assert stale_receive.status_code == 201, stale_receive.text
    stale_batch_id = int(stale_receive.json()["id"])

    depleted = client.patch(
        f"/api/inventory/batches/{stale_batch_id}?force=true",
        json={"quantity": 0},
        headers=auth_headers,
    )
    assert depleted.status_code == 200, depleted.text

    live_receive = client.post(
        "/api/inventory/receive",
        json={
            "item_id": item["id"],
            "batch_no": live_batch_no,
            "quantity": 15,
            "unit": "kg",
            "cost_per_unit": 1,
            "warehouse_id": warehouse["id"],
            "qc_status": "passed",
        },
        headers=auth_headers,
    )
    assert live_receive.status_code == 201, live_receive.text

    depleted_search = client.get(
        f"/api/inventory/stock?group=materials&q={stale_batch_no}&include_total=true",
        headers=auth_headers,
    )
    assert depleted_search.status_code == 200, depleted_search.text
    assert depleted_search.json() == {"rows": [], "total": 0, "page": 1, "page_size": 500}

    live_search = client.get(
        f"/api/inventory/stock?group=materials&q={live_batch_no}&include_total=true",
        headers=auth_headers,
    )
    assert live_search.status_code == 200, live_search.text
    live_body = live_search.json()
    assert live_body["total"] == 1
    assert len(live_body["rows"]) == 1
    assert live_body["rows"][0]["item_id"] == item["id"]
    assert round(float(live_body["rows"][0]["quantity"]), 2) == 15.00


def test_material_inventory_orders_latest_received_batches_first(client, auth_headers):
    suffix = uuid4().hex[:8].upper()
    warehouses_response = client.get("/api/inventory/warehouses", headers=auth_headers)
    assert warehouses_response.status_code == 200, warehouses_response.text
    warehouse = next(row for row in warehouses_response.json() if row["type"] == "fabric_storage")

    created_items = []
    for label in ("OLD", "NEW"):
        item_response = client.post(
            "/api/inventory/items",
            json={
                "sku": f"FAB-SORT-{suffix}-{label}",
                "name": f"Sort fabric {suffix} {label}",
                "category": "fabric",
                "unit": "kg",
                "default_cost": 1,
                "reorder_level": 0,
                "track_batch": True,
                "is_active": True,
            },
            headers=auth_headers,
        )
        assert item_response.status_code == 201, item_response.text
        created_items.append(item_response.json())

    receipts = (
        (created_items[0], "2026-07-01T08:00:00Z", "OLD"),
        (created_items[1], "2026-07-16T08:00:00Z", "NEW"),
    )
    for item, received_date, label in receipts:
        receive_response = client.post(
            "/api/inventory/receive",
            json={
                "item_id": item["id"],
                "batch_no": f"SORT-{suffix}-{label}",
                "quantity": 10,
                "unit": "kg",
                "cost_per_unit": 1,
                "received_date": received_date,
                "warehouse_id": warehouse["id"],
                "qc_status": "passed",
            },
            headers=auth_headers,
        )
        assert receive_response.status_code == 201, receive_response.text

    stock_response = client.get(
        f"/api/inventory/stock?group=materials&q={suffix}&include_total=true",
        headers=auth_headers,
    )
    assert stock_response.status_code == 200, stock_response.text
    assert [row["item_sku"] for row in stock_response.json()["rows"]][:2] == [
        f"FAB-SORT-{suffix}-NEW",
        f"FAB-SORT-{suffix}-OLD",
    ]

    batches_response = client.get(
        f"/api/inventory/batches?group=materials&q={suffix}&include_total=true",
        headers=auth_headers,
    )
    assert batches_response.status_code == 200, batches_response.text
    assert [row["batch_no"] for row in batches_response.json()["rows"]][:2] == [
        f"SORT-{suffix}-NEW",
        f"SORT-{suffix}-OLD",
    ]


def test_material_receive_gramaj_fills_cutting_passport_defaults(client, auth_headers):
    fabric = client.post(
        "/api/inventory/items",
        json={
            "sku": "FAB-GRAMAJ-AUTO",
            "name": "Auto gramaj fabric",
            "category": "fabric",
            "unit": "kg",
            "default_cost": 3,
            "reorder_level": 0,
            "track_batch": True,
            "is_active": True,
        },
        headers=auth_headers,
    )
    assert fabric.status_code == 201, fabric.text
    item_id = fabric.json()["id"]

    model = client.post(
        "/api/models",
        json={
            "code": "GRAMAJ-AUTO-145",
            "name": "Gramaj autofill model",
            "category": "T-shirt",
            "details_json": {
                "general": {
                    "model_no": "GRAMAJ-AUTO",
                    "variant_no": "145",
                    "pattern_no": "PAT-4220",
                },
            },
            "status": "approved",
        },
        headers=auth_headers,
    )
    assert model.status_code == 201, model.text
    model_id = model.json()["id"]

    bom = client.post(
        f"/api/models/{model_id}/bom",
        json={
            "item_id": item_id,
            "quantity_per_piece": 0.4,
            "unit": "kg",
            "waste_percent": 0,
        },
        headers=auth_headers,
    )
    assert bom.status_code == 201, bom.text

    po = client.post(
        "/api/planning/create-branded-production",
        json={
            "production_type": "branded_stock",
            "model_id": model_id,
            "planned_quantity": 600,
            "items": [
                {"model_id": model_id, "color": "white", "size": size, "planned_quantity": 100}
                for size in ["46", "48", "50", "52", "54", "56"]
            ],
        },
        headers=auth_headers,
    )
    assert po.status_code == 201, po.text
    production_order = po.json()

    warehouses = client.get("/api/inventory/warehouses", headers=auth_headers)
    assert warehouses.status_code == 200, warehouses.text
    warehouse_id = next(row["id"] for row in warehouses.json() if row["type"] == "fabric_storage")

    order_batch = client.post(
        "/api/inventory/receive",
        json={
            "item_id": item_id,
            "batch_no": "GRAMAJ-ORDER-001",
            "order_no": production_order["production_no"],
            "old_code": "OLD-4220",
            "width": 1.8,
            "quantity": 40,
            "gsm": 0.145,
            "unit": "kg",
            "cost_per_unit": 3,
            "warehouse_id": warehouse_id,
            "qc_status": "passed",
        },
        headers=auth_headers,
    )
    assert order_batch.status_code == 201, order_batch.text
    assert float(order_batch.json()["gsm"]) == 0.145

    newer_general_batch = client.post(
        "/api/inventory/receive",
        json={
            "item_id": item_id,
            "batch_no": "GRAMAJ-GENERAL-001",
            "quantity": 80,
            "gsm": 0.2,
            "unit": "kg",
            "cost_per_unit": 3,
            "warehouse_id": warehouse_id,
            "qc_status": "passed",
        },
        headers=auth_headers,
    )
    assert newer_general_batch.status_code == 201, newer_general_batch.text

    defaults = client.get(
        f"/api/cutting-passports/material-defaults?production_order_id={production_order['id']}",
        headers=auth_headers,
    )
    assert defaults.status_code == 200, defaults.text
    body = defaults.json()
    assert body["material_item_id"] == item_id
    assert body["order_no"] == production_order["order_no"]
    assert body["model_code"] == "GRAMAJ-AUTO-145"
    assert body["model_no"] == "GRAMAJ-AUTO"
    assert body["variant"] == "145"
    assert body["mold_no"] == "PAT-4220"
    assert body["fabric_type"] == "Auto gramaj fabric"
    assert body["batch_no"] == "GRAMAJ-ORDER-001"
    assert body["lot_no"] == "GRAMAJ-ORDER-001"
    assert body["size_range"] == "46-56"
    assert body["sizes"] == ["46", "48", "50", "52", "54", "56"]
    assert body["size_count"] == 6
    assert float(body["width"]) == 1.8
    assert float(body["gramage"]) == 0.145


def test_cutting_passport_defaults_use_order_batch_without_gramaj(client, auth_headers):
    fabric = client.post(
        "/api/inventory/items",
        json={
            "sku": "FAB-NO-GRAMAJ-AUTO",
            "name": "No gramaj fabric",
            "category": "fabric",
            "unit": "kg",
            "default_cost": 3,
            "reorder_level": 0,
            "track_batch": True,
            "is_active": True,
        },
        headers=auth_headers,
    )
    assert fabric.status_code == 201, fabric.text
    item_id = fabric.json()["id"]

    model = client.post(
        "/api/models",
        json={"code": "NO-GRAMAJ-100", "name": "No gramaj model", "category": "T-shirt", "status": "approved"},
        headers=auth_headers,
    )
    assert model.status_code == 201, model.text
    model_id = model.json()["id"]

    bom = client.post(
        f"/api/models/{model_id}/bom",
        json={"item_id": item_id, "quantity_per_piece": 0.4, "unit": "kg", "waste_percent": 0},
        headers=auth_headers,
    )
    assert bom.status_code == 201, bom.text

    po = client.post(
        "/api/planning/create-branded-production",
        json={
            "production_type": "branded_stock",
            "model_id": model_id,
            "planned_quantity": 100,
            "items": [{"model_id": model_id, "color": "white", "size": "M", "planned_quantity": 100}],
        },
        headers=auth_headers,
    )
    assert po.status_code == 201, po.text
    production_order = po.json()

    warehouses = client.get("/api/inventory/warehouses", headers=auth_headers)
    assert warehouses.status_code == 200, warehouses.text
    warehouse_id = next(row["id"] for row in warehouses.json() if row["type"] == "fabric_storage")

    order_batch = client.post(
        "/api/inventory/receive",
        json={
            "item_id": item_id,
            "batch_no": "NO-GRAMAJ-ORDER-001",
            "order_no": production_order["production_no"],
            "quantity": 40,
            "unit": "kg",
            "cost_per_unit": 3,
            "warehouse_id": warehouse_id,
            "qc_status": "passed",
        },
        headers=auth_headers,
    )
    assert order_batch.status_code == 201, order_batch.text

    defaults = client.get(
        f"/api/cutting-passports/material-defaults?production_order_id={production_order['id']}",
        headers=auth_headers,
    )
    assert defaults.status_code == 200, defaults.text
    body = defaults.json()
    assert body["batch_no"] == "NO-GRAMAJ-ORDER-001"
    assert body["lot_no"] == "NO-GRAMAJ-ORDER-001"
    assert body["fabric_type"] == "No gramaj fabric"
    assert body["sizes"] == ["M"]
    assert body["size_count"] == 1
    assert body["gramage"] is None


def test_admin_can_set_inventory_stock_quantity(client, auth_headers):
    create = client.post(
        "/api/inventory/items",
        json={
            "sku": "FAB-STOCK-SET",
            "name": "Stock editable fabric",
            "category": "fabric",
            "unit": "kg",
            "default_cost": 1,
            "reorder_level": 0,
            "track_batch": True,
            "is_active": True,
        },
        headers=auth_headers,
    )
    assert create.status_code == 201, create.text
    item = create.json()

    warehouses = client.get("/api/inventory/warehouses", headers=auth_headers)
    assert warehouses.status_code == 200, warehouses.text
    warehouse_id = next(row["id"] for row in warehouses.json() if row["type"] == "fabric_storage")

    receive = client.post(
        "/api/inventory/receive",
        json={
            "item_id": item["id"],
            "batch_no": "STOCK-SET-001",
            "quantity": 10,
            "unit": "kg",
            "cost_per_unit": 1,
            "warehouse_id": warehouse_id,
            "qc_status": "passed",
        },
        headers=auth_headers,
    )
    assert receive.status_code == 201, receive.text

    increase = client.patch(
        f"/api/inventory/stock/{item['id']}",
        json={"quantity": 15, "unit": "kg"},
        headers=auth_headers,
    )
    assert increase.status_code == 200, increase.text
    assert increase.json()["previous_quantity"] == 10
    assert increase.json()["quantity"] == 15
    assert increase.json()["delta"] == 5
    assert increase.json()["movement_type"] == "adjustment"

    decrease = client.patch(
        f"/api/inventory/stock/{item['id']}",
        json={"quantity": 8, "unit": "kg"},
        headers=auth_headers,
    )
    assert decrease.status_code == 200, decrease.text
    assert decrease.json()["previous_quantity"] == 15
    assert decrease.json()["quantity"] == 8
    assert decrease.json()["delta"] == -7
    assert decrease.json()["movement_type"] == "issue"

    stock = client.get("/api/inventory/stock?group=materials&q=FAB-STOCK-SET", headers=auth_headers)
    assert stock.status_code == 200, stock.text
    assert len(stock.json()) == 1
    assert stock.json()[0]["quantity"] == 8

    batches = client.get(f"/api/inventory/batches?item_id={item['id']}", headers=auth_headers)
    assert batches.status_code == 200, batches.text
    batch_rows = [row for row in batches.json() if row["batch_no"] == "STOCK-SET-001"]
    assert len(batch_rows) == 1
    assert batch_rows[0]["quantity"] == 8
    assert batch_rows[0]["available_quantity"] == 8


def test_stock_quantity_update_rejects_multi_batch_material(client, auth_headers):
    create = client.post(
        "/api/inventory/items",
        json={
            "sku": "FAB-STOCK-MULTI",
            "name": "Multi batch stock fabric",
            "category": "fabric",
            "unit": "kg",
            "default_cost": 1,
            "reorder_level": 0,
            "track_batch": True,
            "is_active": True,
        },
        headers=auth_headers,
    )
    assert create.status_code == 201, create.text
    item = create.json()

    warehouses = client.get("/api/inventory/warehouses", headers=auth_headers)
    assert warehouses.status_code == 200, warehouses.text
    warehouse_id = next(row["id"] for row in warehouses.json() if row["type"] == "fabric_storage")

    for batch_no, qty in (("STOCK-MULTI-001", 10), ("STOCK-MULTI-002", 12)):
        receive = client.post(
            "/api/inventory/receive",
            json={
                "item_id": item["id"],
                "batch_no": batch_no,
                "quantity": qty,
                "unit": "kg",
                "cost_per_unit": 1,
                "warehouse_id": warehouse_id,
                "qc_status": "passed",
            },
            headers=auth_headers,
        )
        assert receive.status_code == 201, receive.text

    response = client.patch(
        f"/api/inventory/stock/{item['id']}",
        json={"quantity": 5, "unit": "kg"},
        headers=auth_headers,
    )
    assert response.status_code == 409, response.text
    assert "multiple active batches" in response.json()["detail"]

    batches = client.get(f"/api/inventory/batches?item_id={item['id']}", headers=auth_headers)
    assert batches.status_code == 200, batches.text
    quantities = {row["batch_no"]: row["quantity"] for row in batches.json()}
    assert quantities["STOCK-MULTI-001"] == 10
    assert quantities["STOCK-MULTI-002"] == 12


def test_stock_batch_editor_updates_receipt_fields_and_exact_quantity(client, auth_headers):
    item_response = client.post(
        "/api/inventory/items",
        json={
            "sku": "FAB-BATCH-EDIT",
            "name": "Editable batch fabric",
            "category": "fabric",
            "unit": "kg",
            "default_cost": 1,
            "reorder_level": 0,
            "track_batch": True,
            "is_active": True,
        },
        headers=auth_headers,
    )
    assert item_response.status_code == 201, item_response.text
    item = item_response.json()

    warehouses_response = client.get("/api/inventory/warehouses", headers=auth_headers)
    assert warehouses_response.status_code == 200, warehouses_response.text
    warehouses = warehouses_response.json()
    source_warehouse = warehouses[0]
    target_warehouse = warehouses[1]
    suppliers_response = client.get("/api/suppliers", headers=auth_headers)
    assert suppliers_response.status_code == 200, suppliers_response.text
    supplier = suppliers_response.json()[0]

    receive = client.post(
        "/api/inventory/receive",
        json={
            "item_id": item["id"],
            "batch_no": "BATCH-BEFORE",
            "quantity": 10,
            "unit": "kg",
            "cost_per_unit": 1,
            "warehouse_id": source_warehouse["id"],
            "qc_status": "pending",
        },
        headers=auth_headers,
    )
    assert receive.status_code == 201, receive.text
    batch_id = receive.json()["id"]

    update = client.patch(
        f"/api/inventory/batches/{batch_id}",
        json={
            "batch_no": "BATCH-AFTER",
            "supplier_id": supplier["id"],
            "color": "navy",
            "old_code": "OLD-7",
            "color_code": "NV-01",
            "color_status": "approved",
            "order_no": "ORDER-42",
            "width": 1.8,
            "gsm": 0.145,
            "quantity": 7.5,
            "piece_count": 30,
            "processes": "wash, inspect",
            "unit": "kg",
            "cost_per_unit": 6.35,
            "image_url": "/storage/model-files/batch-edit.png",
            "received_date": "2026-07-12T08:30:00Z",
            "warehouse_id": target_warehouse["id"],
            "qc_status": "passed",
        },
        headers=auth_headers,
    )
    assert update.status_code == 200, update.text
    updated = update.json()
    assert updated["batch_no"] == "BATCH-AFTER"
    assert updated["supplier_id"] == supplier["id"]
    assert updated["supplier_name"] == supplier["name"]
    assert updated["warehouse_id"] == target_warehouse["id"]
    assert updated["warehouse_name"] == target_warehouse["name"]
    assert updated["quantity"] == 7.5
    assert updated["available_quantity"] == 7.5
    assert updated["cost_per_unit"] == 6.35
    assert updated["piece_count"] == 30
    assert updated["qc_status"] == "passed"

    batches = client.get(f"/api/inventory/batches?item_id={item['id']}", headers=auth_headers)
    assert batches.status_code == 200, batches.text
    saved = next(row for row in batches.json() if row["id"] == batch_id)
    assert saved["color_code"] == "NV-01"
    assert saved["order_no"] == "ORDER-42"
    assert saved["processes"] == "wash, inspect"


def test_stock_batch_editor_reassigns_an_unused_batch_to_an_existing_material(client, auth_headers):
    suffix = uuid4().hex[:8].upper()

    def create_material(label: str, unit: str = "kg"):
        response = client.post(
            "/api/inventory/items",
            json={
                "sku": f"FAB-BATCH-MOVE-{label}-{suffix}",
                "name": f"Batch move {label} {suffix}",
                "category": "fabric",
                "unit": unit,
                "default_cost": 1,
                "reorder_level": 0,
                "track_batch": True,
                "is_active": True,
            },
            headers=auth_headers,
        )
        assert response.status_code == 201, response.text
        return response.json()

    source_item = create_material("SOURCE")
    target_item = create_material("TARGET")
    wrong_unit_item = create_material("WRONG-UNIT", unit="m")
    warehouses = client.get("/api/inventory/warehouses", headers=auth_headers)
    assert warehouses.status_code == 200, warehouses.text
    receive = client.post(
        "/api/inventory/receive",
        json={
            "item_id": source_item["id"],
            "batch_no": f"MOVE-{suffix}",
            "quantity": 12.5,
            "unit": "kg",
            "cost_per_unit": 2,
            "warehouse_id": warehouses.json()[0]["id"],
            "qc_status": "passed",
        },
        headers=auth_headers,
    )
    assert receive.status_code == 201, receive.text
    batch_id = int(receive.json()["id"])

    wrong_unit = client.patch(
        f"/api/inventory/batches/{batch_id}",
        json={"item_id": wrong_unit_item["id"]},
        headers=auth_headers,
    )
    assert wrong_unit.status_code == 409, wrong_unit.text
    assert "unit" in wrong_unit.json()["detail"].lower()

    update = client.patch(
        f"/api/inventory/batches/{batch_id}",
        json={"item_id": target_item["id"]},
        headers=auth_headers,
    )
    assert update.status_code == 200, update.text
    updated = update.json()
    assert updated["item_id"] == target_item["id"]
    assert updated["item_name"] == target_item["name"]
    assert updated["quantity"] == 12.5

    source = client.get(
        f"/api/inventory/items?group=materials&q={source_item['sku']}",
        headers=auth_headers,
    )
    assert source.status_code == 200, source.text
    assert source.json()[0]["name"] == source_item["name"]
    target_batches = client.get(f"/api/inventory/batches?item_id={target_item['id']}", headers=auth_headers)
    assert target_batches.status_code == 200, target_batches.text
    assert any(row["id"] == batch_id for row in target_batches.json())


def test_reserved_batch_delete_releases_reservation_and_archives_inventory(client, auth_headers):
    from app.db import session as session_module
    from app.models import MaterialReservation, Model, ProductionOrder, StockBatch, StockMovement

    suffix = uuid4().hex[:8].upper()
    materials = []
    for label in ("SOURCE", "TARGET"):
        response = client.post(
            "/api/inventory/items",
            json={
                "sku": f"FAB-RESERVED-MOVE-{label}-{suffix}",
                "name": f"Reserved batch move {label} {suffix}",
                "category": "fabric",
                "unit": "kg",
                "default_cost": 1,
                "reorder_level": 0,
                "track_batch": True,
                "is_active": True,
            },
            headers=auth_headers,
        )
        assert response.status_code == 201, response.text
        materials.append(response.json())

    warehouses = client.get("/api/inventory/warehouses", headers=auth_headers)
    assert warehouses.status_code == 200, warehouses.text
    warehouse_id = int(warehouses.json()[0]["id"])
    receive = client.post(
        "/api/inventory/receive",
        json={
            "item_id": materials[0]["id"],
            "batch_no": f"RESERVED-MOVE-{suffix}",
            "quantity": 10,
            "unit": "kg",
            "cost_per_unit": 1,
            "warehouse_id": warehouse_id,
            "qc_status": "passed",
        },
        headers=auth_headers,
    )
    assert receive.status_code == 201, receive.text
    batch_id = int(receive.json()["id"])

    with session_module.SessionLocal() as db:
        model = Model(
            code=f"RESERVED-MOVE-{suffix}",
            name=f"Reserved move model {suffix}",
            status="approved",
        )
        db.add(model)
        db.flush()
        production_order = ProductionOrder(
            production_no=f"PO-RESERVED-MOVE-{suffix}",
            production_type="branded_stock",
            model_id=model.id,
            fabric_batch_id=batch_id,
            planned_quantity=1,
        )
        db.add(production_order)
        db.flush()
        db.add(MaterialReservation(
            reservation_no=f"MR-RESERVED-MOVE-{suffix}",
            production_order_id=production_order.id,
            item_id=materials[0]["id"],
            stock_batch_id=batch_id,
            warehouse_id=warehouse_id,
            reserved_quantity=3,
            consumed_quantity=0,
            released_quantity=0,
            unit="kg",
            status="reserved",
            reservation_type="material",
            source="manual",
        ))
        db.commit()

    denied_reassignment = client.patch(
        f"/api/inventory/batches/{batch_id}",
        json={"item_id": materials[1]["id"]},
        headers=auth_headers,
    )
    assert denied_reassignment.status_code == 409, denied_reassignment.text

    reassigned = client.patch(
        f"/api/inventory/batches/{batch_id}?force=true",
        json={"item_id": materials[1]["id"]},
        headers=auth_headers,
    )
    assert reassigned.status_code == 200, reassigned.text
    assert reassigned.json()["item_id"] == materials[1]["id"]
    assert reassigned.json()["reserved_quantity"] == 3
    assert reassigned.json()["available_quantity"] == 7

    with session_module.SessionLocal() as db:
        reservation = db.query(MaterialReservation).filter(
            MaterialReservation.stock_batch_id == batch_id,
        ).one()
        assert reservation.item_id == materials[1]["id"]

    deleted = client.delete(f"/api/inventory/batches/{batch_id}", headers=auth_headers)
    assert deleted.status_code == 204, deleted.text
    with session_module.SessionLocal() as db:
        batch = db.get(StockBatch, batch_id)
        assert batch is not None
        assert float(batch.quantity) == 0
        assert batch.qc_status == "hold"
        assert batch.archived_at is not None
        reservation = db.query(MaterialReservation).filter(
            MaterialReservation.stock_batch_id == batch_id,
        ).one()
        assert reservation.status == "released"
        assert float(reservation.released_quantity) == 3
        deletion_issue = db.query(StockMovement).filter(
            StockMovement.batch_id == batch_id,
            StockMovement.reference_type == "StockBatchDelete",
        ).one()
        assert float(deletion_issue.quantity) == 10
        production_order = db.query(ProductionOrder).filter(
            ProductionOrder.production_no == f"PO-RESERVED-MOVE-{suffix}",
        ).one()
        assert production_order.fabric_batch_id == batch_id

    visible = client.get(
        f"/api/inventory/batches?item_id={materials[1]['id']}",
        headers=auth_headers,
    )
    assert visible.status_code == 200, visible.text
    assert all(row["id"] != batch_id for row in visible.json())


def test_stock_batch_editor_rejects_material_change_after_batch_usage(client, auth_headers):
    suffix = uuid4().hex[:8].upper()
    created_items = []
    for label in ("SOURCE", "TARGET"):
        response = client.post(
            "/api/inventory/items",
            json={
                "sku": f"FAB-BATCH-USED-{label}-{suffix}",
                "name": f"Used batch {label} {suffix}",
                "category": "fabric",
                "unit": "kg",
                "default_cost": 1,
                "reorder_level": 0,
                "track_batch": True,
                "is_active": True,
            },
            headers=auth_headers,
        )
        assert response.status_code == 201, response.text
        created_items.append(response.json())

    warehouses = client.get("/api/inventory/warehouses", headers=auth_headers)
    assert warehouses.status_code == 200, warehouses.text
    receive = client.post(
        "/api/inventory/receive",
        json={
            "item_id": created_items[0]["id"],
            "batch_no": f"USED-{suffix}",
            "quantity": 10,
            "unit": "kg",
            "cost_per_unit": 1,
            "warehouse_id": warehouses.json()[0]["id"],
            "qc_status": "passed",
        },
        headers=auth_headers,
    )
    assert receive.status_code == 201, receive.text
    batch_id = int(receive.json()["id"])
    adjustment = client.patch(
        f"/api/inventory/batches/{batch_id}",
        json={"quantity": 9},
        headers=auth_headers,
    )
    assert adjustment.status_code == 200, adjustment.text

    update = client.patch(
        f"/api/inventory/batches/{batch_id}",
        json={"item_id": created_items[1]["id"]},
        headers=auth_headers,
    )
    assert update.status_code == 409, update.text
    assert "reserved or used" in update.json()["detail"]


def test_stock_summary_hides_only_inactive_materials_without_current_stock(client, auth_headers):
    suffix = uuid4().hex[:8].upper()
    created = client.post(
        "/api/inventory/items",
        json={
            "sku": f"FAB-INACTIVE-STOCK-{suffix}",
            "name": f"Inactive stock visibility {suffix}",
            "category": "fabric",
            "unit": "kg",
            "default_cost": 1,
            "reorder_level": 0,
            "track_batch": True,
            "is_active": True,
        },
        headers=auth_headers,
    )
    assert created.status_code == 201, created.text
    item = created.json()
    warehouses = client.get("/api/inventory/warehouses", headers=auth_headers)
    assert warehouses.status_code == 200, warehouses.text
    received = client.post(
        "/api/inventory/receive",
        json={
            "item_id": item["id"],
            "batch_no": f"INACTIVE-STOCK-{suffix}",
            "quantity": 5,
            "unit": "kg",
            "cost_per_unit": 1,
            "warehouse_id": warehouses.json()[0]["id"],
            "qc_status": "passed",
        },
        headers=auth_headers,
    )
    assert received.status_code == 201, received.text
    batch_id = int(received.json()["id"])

    archived = client.patch(
        f"/api/inventory/items/{item['id']}",
        json={**item, "is_active": False},
        headers=auth_headers,
    )
    assert archived.status_code == 200, archived.text
    with_stock = client.get(
        f"/api/inventory/stock?group=materials&q={item['sku']}&include_total=true",
        headers=auth_headers,
    )
    assert with_stock.status_code == 200, with_stock.text
    assert with_stock.json()["total"] == 1

    zeroed = client.patch(
        f"/api/inventory/batches/{batch_id}?force=true",
        json={"quantity": 0},
        headers=auth_headers,
    )
    assert zeroed.status_code == 200, zeroed.text
    empty = client.get(
        f"/api/inventory/stock?group=materials&q={item['sku']}&include_total=true",
        headers=auth_headers,
    )
    assert empty.status_code == 200, empty.text
    assert empty.json() == {"rows": [], "total": 0, "page": 1, "page_size": 500}


def test_material_inventory_report_exports_grouped_counts_kg_and_grand_total(client, auth_headers):
    suffix = uuid4().hex[:8].upper()
    item_response = client.post(
        "/api/inventory/items",
        json={
            "sku": f"FAB-REPORT-{suffix}",
            "name": f"Report fabric {suffix}",
            "category": "fabric",
            "unit": "kg",
            "default_cost": 1,
            "reorder_level": 0,
            "track_batch": True,
            "is_active": True,
        },
        headers=auth_headers,
    )
    assert item_response.status_code == 201, item_response.text
    item = item_response.json()
    warehouses = client.get("/api/inventory/warehouses", headers=auth_headers)
    assert warehouses.status_code == 200, warehouses.text
    warehouse_id = warehouses.json()[0]["id"]
    for batch_no, quantity, piece_count in (("REPORT-A", 5.25, 3), ("REPORT-B", 4.75, 4)):
        receive = client.post(
            "/api/inventory/receive",
            json={
                "item_id": item["id"],
                "batch_no": f"{batch_no}-{suffix}",
                "quantity": quantity,
                "piece_count": piece_count,
                "unit": "kg",
                "cost_per_unit": 1,
                "warehouse_id": warehouse_id,
                "qc_status": "passed",
            },
            headers=auth_headers,
        )
        assert receive.status_code == 201, receive.text

    excel_response = client.get(
        "/api/inventory/reports/material-stock.xlsx?lang=en",
        headers=auth_headers,
    )
    assert excel_response.status_code == 200, excel_response.text
    assert excel_response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert ".xlsx" in excel_response.headers["content-disposition"]
    workbook = load_workbook(BytesIO(excel_response.content), data_only=False)
    sheet = workbook["Material Inventory"]
    material_row = next(
        row
        for row in sheet.iter_rows(min_row=6, values_only=True)
        if row[1] == item["name"]
    )
    assert material_row[3] == 2
    assert material_row[4] == 7
    assert material_row[5] == 10
    grand_total_row = sheet.max_row
    assert sheet.cell(grand_total_row, 1).value == "Grand total"
    assert str(sheet.cell(grand_total_row, 4).value).startswith("=SUM(")
    assert str(sheet.cell(grand_total_row, 5).value).startswith("=SUM(")
    assert str(sheet.cell(grand_total_row, 6).value).startswith("=SUM(")

    pdf_response = client.get(
        "/api/inventory/reports/material-stock.pdf?lang=en",
        headers=auth_headers,
    )
    assert pdf_response.status_code == 200, pdf_response.text
    assert pdf_response.headers["content-type"].startswith("application/pdf")
    assert ".pdf" in pdf_response.headers["content-disposition"]
    assert pdf_response.content.startswith(b"%PDF")
    assert len(pdf_response.content) > 10_000


def test_admin_can_delete_unused_stock_receipt_batch(client, auth_headers):
    suffix = uuid4().hex[:8].upper()
    item_response = client.post(
        "/api/inventory/items",
        json={
            "sku": f"FAB-BATCH-DELETE-{suffix}",
            "name": f"Duplicate receipt {suffix}",
            "category": "fabric",
            "unit": "kg",
            "default_cost": 1,
            "reorder_level": 0,
            "track_batch": True,
            "is_active": True,
        },
        headers=auth_headers,
    )
    assert item_response.status_code == 201, item_response.text
    warehouses = client.get("/api/inventory/warehouses", headers=auth_headers)
    assert warehouses.status_code == 200, warehouses.text
    receive = client.post(
        "/api/inventory/receive",
        json={
            "item_id": item_response.json()["id"],
            "batch_no": f"DUP-{suffix}",
            "quantity": 10,
            "unit": "kg",
            "cost_per_unit": 1,
            "warehouse_id": warehouses.json()[0]["id"],
            "qc_status": "passed",
        },
        headers=auth_headers,
    )
    assert receive.status_code == 201, receive.text
    batch_id = int(receive.json()["id"])

    deleted = client.delete(f"/api/inventory/batches/{batch_id}", headers=auth_headers)
    assert deleted.status_code == 204, deleted.text
    batches = client.get(
        f"/api/inventory/batches?item_id={item_response.json()['id']}",
        headers=auth_headers,
    )
    assert batches.status_code == 200, batches.text
    assert all(int(row["id"]) != batch_id for row in batches.json())


def test_stock_batch_delete_lock_does_not_lock_outer_join():
    sql = str(
        _locked_stock_batch_statement(90).compile(
            dialect=postgresql.dialect(),
            compile_kwargs={"literal_binds": True},
        )
    ).upper()

    assert "FOR UPDATE" in sql
    assert " JOIN " not in sql


def test_stock_batch_delete_reservation_lock_targets_only_reservation_rows():
    sql = str(
        _locked_active_batch_reservations_statement(90).compile(
            dialect=postgresql.dialect(),
            compile_kwargs={"literal_binds": True},
        )
    ).upper()

    assert "FOR UPDATE OF MATERIAL_RESERVATIONS" in sql
    assert " JOIN " not in sql


def test_stock_batch_delete_flushes_receipts_before_parent_batch():
    calls: list[object] = []

    class FakeSession:
        def delete(self, row):
            calls.append(row)

        def flush(self):
            calls.append("flush")

    first = object()
    second = object()
    _delete_stock_batch_receipt_movements(FakeSession(), [first, second])

    assert calls == [first, second, "flush"]


def test_stock_batch_delete_archives_used_batch_and_reduces_remaining_inventory(client, auth_headers):
    from app.db import session as session_module
    from app.models import StockBatch, StockMovement

    suffix = uuid4().hex[:8].upper()
    item_response = client.post(
        "/api/inventory/items",
        json={
            "sku": f"FAB-BATCH-USED-{suffix}",
            "name": f"Used receipt {suffix}",
            "category": "fabric",
            "unit": "kg",
            "default_cost": 1,
            "reorder_level": 0,
            "track_batch": True,
            "is_active": True,
        },
        headers=auth_headers,
    )
    assert item_response.status_code == 201, item_response.text
    warehouses = client.get("/api/inventory/warehouses", headers=auth_headers)
    assert warehouses.status_code == 200, warehouses.text
    receive = client.post(
        "/api/inventory/receive",
        json={
            "item_id": item_response.json()["id"],
            "batch_no": f"USED-{suffix}",
            "quantity": 10,
            "unit": "kg",
            "cost_per_unit": 1,
            "warehouse_id": warehouses.json()[0]["id"],
            "qc_status": "passed",
        },
        headers=auth_headers,
    )
    assert receive.status_code == 201, receive.text
    batch_id = int(receive.json()["id"])
    adjusted = client.patch(
        f"/api/inventory/batches/{batch_id}",
        json={"quantity": 9},
        headers=auth_headers,
    )
    assert adjusted.status_code == 200, adjusted.text

    deleted = client.delete(f"/api/inventory/batches/{batch_id}", headers=auth_headers)
    assert deleted.status_code == 204, deleted.text

    with session_module.SessionLocal() as db:
        batch = db.get(StockBatch, batch_id)
        assert batch is not None
        assert float(batch.quantity) == 0
        assert batch.qc_status == "hold"
        assert batch.archived_at is not None
        deletion_issue = db.query(StockMovement).filter(
            StockMovement.batch_id == batch_id,
            StockMovement.reference_type == "StockBatchDelete",
        ).one()
        assert float(deletion_issue.quantity) == 9

    visible = client.get(
        f"/api/inventory/batches?item_id={item_response.json()['id']}",
        headers=auth_headers,
    )
    assert visible.status_code == 200, visible.text
    assert all(row["id"] != batch_id for row in visible.json())


def test_admin_can_create_update_and_delete_supplier(client, auth_headers):
    create = client.post(
        "/api/suppliers",
        json={
            "name": "Delete Me Supplier",
            "phone": "+998900000001",
            "email": "delete-me@example.com",
            "address": "Tashkent",
            "notes": "temporary",
        },
        headers=auth_headers,
    )
    assert create.status_code == 201, create.text
    supplier_id = create.json()["id"]

    update = client.patch(
        f"/api/suppliers/{supplier_id}",
        json={
            "name": "Edited Supplier",
            "phone": "+998900000002",
            "email": "edited@example.com",
            "address": "Samarkand",
            "notes": "edited",
        },
        headers=auth_headers,
    )
    assert update.status_code == 200, update.text
    assert update.json()["name"] == "Edited Supplier"

    delete = client.delete(f"/api/suppliers/{supplier_id}", headers=auth_headers)
    assert delete.status_code == 204, delete.text


def test_cannot_delete_supplier_linked_to_stock(client, auth_headers):
    supplier = client.post(
        "/api/suppliers",
        json={"name": "Linked Supplier"},
        headers=auth_headers,
    )
    assert supplier.status_code == 201, supplier.text
    supplier_id = supplier.json()["id"]

    items = client.get("/api/inventory/items?group=materials", headers=auth_headers)
    assert items.status_code == 200, items.text
    item_id = items.json()[0]["id"]

    warehouses = client.get("/api/inventory/warehouses", headers=auth_headers)
    assert warehouses.status_code == 200, warehouses.text
    warehouse_id = next(row["id"] for row in warehouses.json() if row["type"] == "fabric_storage")

    receive = client.post(
        "/api/inventory/receive",
        json={
            "item_id": item_id,
            "batch_no": "LINKED-SUPPLIER-001",
            "supplier_id": supplier_id,
            "quantity": 1,
            "unit": "kg",
            "cost_per_unit": 1,
            "warehouse_id": warehouse_id,
            "qc_status": "passed",
        },
        headers=auth_headers,
    )
    assert receive.status_code == 201, receive.text

    delete = client.delete(f"/api/suppliers/{supplier_id}", headers=auth_headers)
    assert delete.status_code == 409, delete.text
