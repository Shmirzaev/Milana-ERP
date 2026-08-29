from datetime import date, datetime, timedelta, timezone
from io import BytesIO
from types import SimpleNamespace
from uuid import uuid4

from openpyxl import load_workbook
import pytest

from app.api.routes.sewing_daily_reports import _report_model_info
from app.core.security import create_access_token
from app.db.session import SessionLocal
from app.models import AuditLog


def _create_report_test_flow(client, headers, capacity: int = 500, factory_code: str | None = None) -> dict:
    suffix = uuid4().hex[:8].upper()
    created = client.post(
        "/api/sewing-flows",
        json={
            "name": f"Daily Report Test {suffix}",
            "code": f"DR-{suffix}",
            "capacity_per_day": capacity,
            "is_active": True,
            **({"factory_code": factory_code} if factory_code else {}),
        },
        headers=headers,
    )
    assert created.status_code == 201, created.text
    return created.json()


def test_sewing_daily_report_saves_without_mutating_workflow(client, auth_headers):
    created = client.post(
        "/api/planning/create-branded-production",
        json={
            "production_type": "branded_stock",
            "model_id": 1,
            "planned_quantity": 120,
            "items": [{"model_id": 1, "color": "white", "size": "M", "planned_quantity": 120}],
        },
        headers=auth_headers,
    )
    assert created.status_code == 201, created.text
    po_id = int(created.json()["id"])
    passport_no = f"KR-DAILY-{po_id}"
    passport = client.post(
        "/api/cutting-passports",
        json={
            "passport_no": passport_no,
            "date": datetime.now(timezone.utc).isoformat(),
            "production_order_id": po_id,
            "has_print": False,
        },
        headers=auth_headers,
    )
    assert passport.status_code == 201, passport.text

    work_orders = client.get(f"/api/work-orders?production_order_id={po_id}", headers=auth_headers)
    assert work_orders.status_code == 200, work_orders.text
    sewing_wo = next(row for row in work_orders.json() if row["operation"] == "sewing")

    flow = _create_report_test_flow(client, auth_headers)

    assignment = client.post(
        f"/api/work-orders/{sewing_wo['id']}/assignments",
        json={
            "work_order_id": sewing_wo["id"],
            "sewing_flow_id": flow["id"],
            "quantity": 120,
            "planned_start": datetime.now(timezone.utc).isoformat(),
            "planned_end": (datetime.now(timezone.utc) + timedelta(days=1)).isoformat(),
        },
        headers=auth_headers,
    )
    assert assignment.status_code == 201, assignment.text
    assignment_id = int(assignment.json()["id"])

    context = client.get(
        f"/api/sewing-daily-reports/line-context?sewing_flow_id={flow['id']}",
        headers=auth_headers,
    )
    assert context.status_code == 200, context.text
    active = context.json()["active_work_orders"]
    active_row = next(
        row
        for row in active
        if row["work_order_id"] == sewing_wo["id"] and row["sewing_assignment_id"] == assignment_id
    )
    assert active_row["model_id"] == 1
    assert active_row["model_no"]
    assert "variant_no" in active_row
    assert "model_image_url" in active_row
    assert "fabric_image_url" in active_row
    assert active_row["kroy_no"] == passport_no

    report_date = date.today().isoformat()
    saved = client.post(
        "/api/sewing-daily-reports",
        json={
            "report_date": report_date,
            "sewing_flow_id": flow["id"],
            "work_order_id": sewing_wo["id"],
            "sewing_assignment_id": assignment_id,
            "sewn_qty": 42,
            "defective_qty": 2,
            "defect_reason": "Needle mark",
        },
        headers=auth_headers,
    )
    assert saved.status_code == 201, saved.text
    body = saved.json()
    assert body["line_code"] == flow["code"]
    assert body["sewn_qty"] == 42
    assert body["defective_qty"] == 2
    assert body["manual_model_no"] is None
    assert body["manual_variant_no"] is None
    assert body["model_id"] == active_row["model_id"]
    assert body["model_no"] == active_row["model_no"]
    assert body["variant_no"] == active_row["variant_no"]
    assert body["fabric_image_url"] == active_row["fabric_image_url"]
    assert body["kroy_no"] == passport_no

    listed = client.get(f"/api/sewing-daily-reports?report_date={report_date}", headers=auth_headers)
    assert listed.status_code == 200, listed.text
    listed_body = listed.json()
    assert listed_body["total_sewn_qty"] == 42
    assert listed_body["total_defective_qty"] == 2
    assert listed_body["summary"][0]["line_code"] == flow["code"]
    assert listed_body["summary"][0]["orders"]
    assert listed_body["summary"][0]["models"][0]["model_id"] == 1
    assert listed_body["summary"][0]["models"][0]["model_no"] == active_row["model_no"]
    assert listed_body["summary"][0]["models"][0]["variant_no"] == active_row["variant_no"]
    assert listed_body["summary"][0]["kroy_nos"] == [passport_no]
    assert listed_body["rows"][0]["model_id"] == 1
    assert listed_body["rows"][0]["model_no"] == active_row["model_no"]
    assert listed_body["rows"][0]["variant_no"] == active_row["variant_no"]
    assert listed_body["rows"][0]["kroy_no"] == passport_no

    mixed_identity = client.post(
        "/api/sewing-daily-reports",
        json={
            "report_date": report_date,
            "sewing_flow_id": flow["id"],
            "work_order_id": sewing_wo["id"],
            "sewing_assignment_id": assignment_id,
            "manual_model_no": "MANUAL-7381",
            "manual_variant_no": "V-4",
            "sewn_qty": 10,
            "defective_qty": 0,
        },
        headers=auth_headers,
    )
    assert mixed_identity.status_code == 422, mixed_identity.text

    converted_to_manual = client.patch(
        f"/api/sewing-daily-reports/{body['id']}",
        json={
            "report_date": report_date,
            "manual_model_no": "MANUAL-7381",
            "manual_variant_no": None,
            "kroy_no": "KR-MANUAL-7381",
            "sewn_qty": 42,
            "defective_qty": 2,
            "defect_reason": "Needle mark",
        },
        headers=auth_headers,
    )
    assert converted_to_manual.status_code == 200, converted_to_manual.text
    manual_body = converted_to_manual.json()
    assert manual_body["work_order_id"] is None
    assert manual_body["sewing_assignment_id"] is None
    assert manual_body["production_order_id"] is None
    assert manual_body["production_batch_id"] is None
    assert manual_body["model_id"] is None
    assert manual_body["model_no"] == "MANUAL-7381"
    assert manual_body["variant_no"] is None
    assert manual_body["model_image_url"] is None
    assert manual_body["fabric_image_url"] is None

    unchanged_work_order = client.get(f"/api/work-orders/{sewing_wo['id']}", headers=auth_headers)
    assert unchanged_work_order.status_code == 200, unchanged_work_order.text
    assert unchanged_work_order.json()["passed_qty"] == sewing_wo["passed_qty"]
    assert unchanged_work_order.json()["failed_qty"] == sewing_wo["failed_qty"]

    assignments_after = client.get(f"/api/work-orders/{sewing_wo['id']}/assignments", headers=auth_headers)
    assert assignments_after.status_code == 200, assignments_after.text
    refreshed_assignment = next(row for row in assignments_after.json() if row["id"] == assignment_id)
    assert refreshed_assignment["completed_qty"] == 0
    assert refreshed_assignment["status"] == "planned"


def test_manual_report_identity_never_borrows_attached_model_media():
    report = SimpleNamespace(
        manual_model_no="MANUAL-ONLY",
        manual_variant_no="",
    )
    payload = _report_model_info(
        db=None,
        report=report,
        production_order=SimpleNamespace(id=999, model_id=1),
    )

    assert payload == {
        "model_id": None,
        "model_code": None,
        "model_no": "MANUAL-ONLY",
        "variant_no": None,
        "model_name": None,
        "model_image_url": None,
        "fabric_image_url": None,
    }


def test_sewing_daily_report_requires_reason_for_defects(client, auth_headers):
    created = client.post(
        "/api/planning/create-branded-production",
        json={
            "production_type": "branded_stock",
            "model_id": 1,
            "planned_quantity": 20,
            "items": [{"model_id": 1, "color": "white", "size": "M", "planned_quantity": 20}],
        },
        headers=auth_headers,
    )
    assert created.status_code == 201, created.text
    po_id = int(created.json()["id"])

    work_orders = client.get(f"/api/work-orders?production_order_id={po_id}", headers=auth_headers)
    assert work_orders.status_code == 200, work_orders.text
    sewing_wo = next(row for row in work_orders.json() if row["operation"] == "sewing")

    flow = _create_report_test_flow(client, auth_headers)
    assigned = client.patch(
        f"/api/work-orders/{sewing_wo['id']}",
        json={"sewing_flow_id": flow["id"]},
        headers=auth_headers,
    )
    assert assigned.status_code == 200, assigned.text

    missing_reason = client.post(
        "/api/sewing-daily-reports",
        json={
            "report_date": date.today().isoformat(),
            "sewing_flow_id": flow["id"],
            "work_order_id": sewing_wo["id"],
            "sewn_qty": 10,
            "defective_qty": 1,
        },
        headers=auth_headers,
    )
    assert missing_reason.status_code == 422, missing_reason.text


def test_sewing_daily_report_allows_manual_entry_without_order(client, auth_headers):
    flow = _create_report_test_flow(client, auth_headers)
    report_date = date.today().isoformat()

    saved = client.post(
        "/api/sewing-daily-reports",
        json={
            "report_date": report_date,
            "sewing_flow_id": flow["id"],
            "work_order_id": None,
            "manual_model_no": "MANUAL-2048",
            "manual_variant_no": "V-17",
            "kroy_no": "KR-MANUAL-17",
            "sewn_qty": 36,
            "defective_qty": 1,
            "defect_reason": "Needle mark",
            "notes": "Entered without an attached order",
        },
        headers=auth_headers,
    )
    assert saved.status_code == 201, saved.text
    body = saved.json()
    assert body["work_order_id"] is None
    assert body["sewing_assignment_id"] is None
    assert body["production_order_id"] is None
    assert body["order_no"] is None
    assert body["model_id"] is None
    assert body["model_no"] == "MANUAL-2048"
    assert body["variant_no"] == "V-17"
    assert body["kroy_no"] == "KR-MANUAL-17"
    assert body["sewn_qty"] == 36

    listed = client.get(f"/api/sewing-daily-reports?report_date={report_date}", headers=auth_headers)
    assert listed.status_code == 200, listed.text
    row = next(item for item in listed.json()["rows"] if item["id"] == body["id"])
    assert row["work_order_id"] is None
    assert row["model_no"] == "MANUAL-2048"
    assert row["kroy_no"] == "KR-MANUAL-17"
    summary = next(item for item in listed.json()["summary"] if item["sewing_flow_id"] == flow["id"])
    assert summary["orders"] == []
    assert summary["models"][0]["model_no"] == "MANUAL-2048"
    assert summary["kroy_nos"] == ["KR-MANUAL-17"]

    missing_model = client.post(
        "/api/sewing-daily-reports",
        json={
            "report_date": report_date,
            "sewing_flow_id": flow["id"],
            "work_order_id": None,
            "sewn_qty": 10,
        },
        headers=auth_headers,
    )
    assert missing_model.status_code == 422, missing_model.text


def test_sewing_daily_report_entry_can_be_corrected_with_audit_history(client, auth_headers):
    flow = _create_report_test_flow(client, auth_headers)
    original_date = "2099-12-20"
    corrected_date = "2099-12-21"
    created = client.post(
        "/api/sewing-daily-reports",
        json={
            "report_date": original_date,
            "sewing_flow_id": flow["id"],
            "work_order_id": None,
            "manual_model_no": "WRONG-100",
            "manual_variant_no": "V-1",
            "kroy_no": "KR-WRONG",
            "sewn_qty": 30,
            "defective_qty": 2,
            "defect_reason": "Needle mark",
            "notes": "Original entry",
        },
        headers=auth_headers,
    )
    assert created.status_code == 201, created.text
    report_id = int(created.json()["id"])

    corrected = client.patch(
        f"/api/sewing-daily-reports/{report_id}",
        json={
            "report_date": corrected_date,
            "manual_model_no": "CORRECT-200",
            "manual_variant_no": "V-2",
            "kroy_no": "KR-CORRECT",
            "sewn_qty": 36,
            "section_quantities": None,
            "section_no": None,
            "section_name": None,
            "top_qty": None,
            "bottom_qty": None,
            "defective_qty": 1,
            "defect_reason": "Hole present",
            "notes": "Corrected entry",
        },
        headers=auth_headers,
    )
    assert corrected.status_code == 200, corrected.text
    body = corrected.json()
    assert body["report_date"] == corrected_date
    assert body["model_no"] == "CORRECT-200"
    assert body["variant_no"] == "V-2"
    assert body["kroy_no"] == "KR-CORRECT"
    assert body["sewn_qty"] == 36
    assert body["defective_qty"] == 1
    assert body["defect_reason"] == "Hole present"
    assert body["notes"] == "Corrected entry"

    old_day = client.get(
        f"/api/sewing-daily-reports?report_date={original_date}",
        headers=auth_headers,
    )
    assert old_day.status_code == 200, old_day.text
    assert all(row["id"] != report_id for row in old_day.json()["rows"])
    corrected_day = client.get(
        f"/api/sewing-daily-reports?report_date={corrected_date}",
        headers=auth_headers,
    )
    assert corrected_day.status_code == 200, corrected_day.text
    assert next(row for row in corrected_day.json()["rows"] if row["id"] == report_id)["sewn_qty"] == 36

    db = SessionLocal()
    try:
        audit = (
            db.query(AuditLog)
            .filter(
                AuditLog.action == "update",
                AuditLog.entity_type == "SewingDailyReport",
                AuditLog.entity_id == report_id,
            )
            .order_by(AuditLog.id.desc())
            .first()
        )
        assert audit is not None
        assert audit.old_value_json["sewn_qty"] == 30
        assert audit.new_value_json["sewn_qty"] == 36
        assert audit.old_value_json["manual_model_no"] == "WRONG-100"
        assert audit.new_value_json["manual_model_no"] == "CORRECT-200"
    finally:
        db.close()

    missing_model = client.patch(
        f"/api/sewing-daily-reports/{report_id}",
        json={
            "report_date": corrected_date,
            "manual_model_no": None,
            "manual_variant_no": None,
            "kroy_no": "KR-CORRECT",
            "sewn_qty": 36,
            "defective_qty": 0,
        },
        headers=auth_headers,
    )
    assert missing_model.status_code == 400, missing_model.text


def test_sewing_daily_report_supports_two_part_dynamic_sections(client, auth_headers):
    flows = client.get("/api/sewing-flows", headers=auth_headers)
    assert flows.status_code == 200, flows.text
    sectioned_flow = next(flow for flow in flows.json() if flow["code"] == "SEW-01")
    report_date = "2099-12-30"

    saved = client.post(
        "/api/sewing-daily-reports",
        json={
            "report_date": report_date,
            "sewing_flow_id": sectioned_flow["id"],
            "work_order_id": None,
            "manual_model_no": "SET-2048",
            "manual_variant_no": "V-17",
            "section_no": 4,
            "top_qty": 18,
            "bottom_qty": 10,
            "sewn_qty": 28,
            "defective_qty": 0,
        },
        headers=auth_headers,
    )
    assert saved.status_code == 201, saved.text
    body = saved.json()
    assert body["section_no"] == 4
    assert body["top_qty"] == 18
    assert body["bottom_qty"] == 10

    listed = client.get(f"/api/sewing-daily-reports?report_date={report_date}", headers=auth_headers)
    assert listed.status_code == 200, listed.text
    row = next(item for item in listed.json()["rows"] if item["id"] == body["id"])
    assert row["section_no"] == 4
    assert row["top_qty"] == 18
    assert row["bottom_qty"] == 10


@pytest.mark.parametrize(
    ("factory_code", "model_no", "report_date"),
    [
        ("BST", "BST-SET-2048", "2099-12-27"),
        ("ECO", "ECO-SET-2048", "2099-12-28"),
    ],
)
def test_factory_sewing_daily_report_supports_sections_and_two_part_quantities(
    client,
    factory_code,
    model_no,
    report_date,
):
    factory_headers = {
        "Authorization": f"Bearer {create_access_token(1, extra={'factory_code': factory_code})}"
    }
    flow = _create_report_test_flow(client, factory_headers, factory_code=factory_code)

    saved = client.post(
        "/api/sewing-daily-reports",
        json={
            "report_date": report_date,
            "sewing_flow_id": flow["id"],
            "work_order_id": None,
            "manual_model_no": model_no,
            "manual_variant_no": "GREY",
            "section_no": 5,
            "top_qty": 24,
            "bottom_qty": 16,
            "sewn_qty": 40,
            "defective_qty": 0,
        },
        headers=factory_headers,
    )
    assert saved.status_code == 201, saved.text
    body = saved.json()
    assert body["line_code"] == flow["code"]
    assert body["section_no"] == 5
    assert body["top_qty"] == 24
    assert body["bottom_qty"] == 16
    assert body["sewn_qty"] == 40

    listed = client.get(
        f"/api/sewing-daily-reports?report_date={report_date}&factory_code={factory_code}",
        headers=factory_headers,
    )
    assert listed.status_code == 200, listed.text
    row = next(item for item in listed.json()["rows"] if item["id"] == body["id"])
    assert row["section_no"] == 5
    assert row["top_qty"] == 24
    assert row["bottom_qty"] == 16


def test_sewing_daily_report_exports_saved_rows_for_chosen_dates(client, auth_headers):
    flow = _create_report_test_flow(client, auth_headers)
    report_date = "2099-12-29"
    saved = client.post(
        "/api/sewing-daily-reports",
        json={
            "report_date": report_date,
            "sewing_flow_id": flow["id"],
            "work_order_id": None,
            "manual_model_no": "EXPORT-2048",
            "manual_variant_no": "V-29",
            "kroy_no": "KR-EXPORT-29",
            "sewn_qty": 48,
            "defective_qty": 2,
            "defect_reason": "hole_present",
            "notes": "Export verification row",
        },
        headers=auth_headers,
    )
    assert saved.status_code == 201, saved.text

    excel_response = client.get(
        (
            "/api/sewing-daily-reports/export.xlsx"
            f"?from_date={report_date}&to_date={report_date}&lang=en"
        ),
        headers=auth_headers,
    )
    assert excel_response.status_code == 200, excel_response.text
    assert excel_response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert ".xlsx" in excel_response.headers["content-disposition"]
    workbook = load_workbook(BytesIO(excel_response.content), data_only=False)
    assert workbook.sheetnames == ["Summary", "Entries"]
    summary_sheet = workbook["Summary"]
    entries_sheet = workbook["Entries"]
    assert summary_sheet["A1"].value == "Daily Sewing Report"
    assert summary_sheet["A7"].value == f"{flow['name']} ({flow['code']})"
    assert summary_sheet["B7"].value == 48
    assert summary_sheet["C7"].value == 2
    assert str(summary_sheet.cell(summary_sheet.max_row, 2).value).startswith("=SUM(")
    assert entries_sheet["G6"].value == "EXPORT-2048"
    assert entries_sheet["H6"].value == "V-29"
    assert entries_sheet["I6"].value == "KR-EXPORT-29"
    assert entries_sheet["J6"].value == 48
    assert entries_sheet["K6"].value == 2
    assert entries_sheet["L6"].value == "Hole present"
    assert entries_sheet["M6"].value == "Export verification row"

    pdf_response = client.get(
        (
            "/api/sewing-daily-reports/export.pdf"
            f"?from_date={report_date}&to_date={report_date}&lang=en"
        ),
        headers=auth_headers,
    )
    assert pdf_response.status_code == 200, pdf_response.text
    assert pdf_response.headers["content-type"].startswith("application/pdf")
    assert ".pdf" in pdf_response.headers["content-disposition"]
    assert pdf_response.content.startswith(b"%PDF")
    assert len(pdf_response.content) > 10_000

    invalid_range = client.get(
        "/api/sewing-daily-reports/export.xlsx?from_date=2099-12-30&to_date=2099-12-29",
        headers=auth_headers,
    )
    assert invalid_range.status_code == 400
