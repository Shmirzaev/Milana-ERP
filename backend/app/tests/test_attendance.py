from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO

from openpyxl import load_workbook
from PIL import Image

from app.models import AttendanceDevice, AttendancePerson, Employee
from app.tests.conftest import TestSessionLocal


INTEGRATION_HEADERS = {"X-Attendance-Token": "test-attendance-token"}


def snapshot(*people, device_key="main-turnstile", source_host="10.100.50.41"):
    return {
        "device": {
            "device_key": device_key,
            "name": device_key.replace("-", " ").title(),
            "vendor": "Hikvision",
            "model": "DS-test",
            "serial_no": "serial-test",
            "source_host": source_host,
            "reported_person_count": len(people),
        },
        "people": list(people),
        "full_snapshot": True,
    }


def person(employee_no: str, name: str):
    return {
        "external_person_id": employee_no,
        "full_name": name,
        "user_type": "normal",
        "is_valid": True,
        "has_face": True,
        "card_count": 0,
        "fingerprint_count": 0,
    }


def test_attendance_snapshot_is_isolated_from_hr(client, auth_headers):
    with TestSessionLocal() as db:
        employee_count_before = db.query(Employee).count()

    response = client.post(
        "/api/attendance/integration/people",
        headers=INTEGRATION_HEADERS,
        json=snapshot(person("735", "Test Person One"), person("371", "Test Person Two")),
    )
    assert response.status_code == 200, response.text
    assert response.json()["created"] == 2

    overview = client.get("/api/attendance/overview?day=2026-08-17", headers=auth_headers)
    assert overview.status_code == 200, overview.text
    assert overview.json()["summary"]["total_people"] == 2
    assert overview.json()["devices"][0]["read_only"] is True

    with TestSessionLocal() as db:
        assert db.query(Employee).count() == employee_count_before
        assert db.query(AttendancePerson).count() == 2


def test_attendance_integration_rejects_invalid_token(client):
    response = client.post(
        "/api/attendance/integration/people",
        headers={"X-Attendance-Token": "wrong"},
        json=snapshot(person("735", "Test Person")),
    )
    assert response.status_code == 401


def test_events_are_idempotent_and_drive_daily_usage(client, auth_headers):
    client.post(
        "/api/attendance/integration/people",
        headers=INTEGRATION_HEADERS,
        json=snapshot(person("735", "Test Person")),
    )
    event = {
        "event_uid": "stable-event-1",
        "external_person_id": "735",
        "occurred_at": "2026-08-17T08:15:00+05:00",
        "direction": "entry",
        "verification_mode": "face",
        "result": "success",
        "door_no": 1,
        "reader_no": 1,
        "serial_no": 10,
    }
    payload = {"device": snapshot()["device"], "events": [event]}
    first = client.post("/api/attendance/integration/events", headers=INTEGRATION_HEADERS, json=payload)
    second = client.post("/api/attendance/integration/events", headers=INTEGRATION_HEADERS, json=payload)
    assert first.json()["inserted"] == 1
    assert second.json()["duplicates"] == 1

    overview = client.get("/api/attendance/overview?day=2026-08-17", headers=auth_headers).json()
    assert overview["summary"]["used_today"] == 1
    assert overview["summary"]["events_today"] == 1
    assert overview["people"][0]["arrival_at"] is not None
    assert overview["people"][0]["departure_at"] is None
    assert overview["people"][0]["worked_minutes"] is None
    assert overview["people"][0]["attendance_status"] == "single_scan"


def test_daily_attendance_uses_first_arrival_and_last_departure_and_exports_report(client, auth_headers):
    client.post(
        "/api/attendance/integration/people",
        headers=INTEGRATION_HEADERS,
        json=snapshot(person("735", "Test Person")),
    )
    events = [
        {
            "event_uid": f"daily-event-{index}",
            "external_person_id": "735",
            "occurred_at": occurred_at,
            "direction": direction,
            "verification_mode": "face",
            "result": "success",
        }
        for index, (occurred_at, direction) in enumerate([
            ("2026-08-17T08:15:00+05:00", "entry"),
            ("2026-08-17T12:00:00+05:00", "unknown"),
            ("2026-08-17T18:20:00+05:00", "exit"),
        ], start=1)
    ]
    imported = client.post(
        "/api/attendance/integration/events",
        headers=INTEGRATION_HEADERS,
        json={"device": snapshot()["device"], "events": events},
    )
    assert imported.status_code == 200, imported.text

    overview = client.get("/api/attendance/overview?day=2026-08-17", headers=auth_headers)
    assert overview.status_code == 200, overview.text
    row = overview.json()["people"][0]
    assert datetime.fromisoformat(row["arrival_at"].replace("Z", "+00:00")) == datetime(
        2026, 8, 17, 3, 15, tzinfo=timezone.utc
    )
    assert datetime.fromisoformat(row["departure_at"].replace("Z", "+00:00")) == datetime(
        2026, 8, 17, 13, 20, tzinfo=timezone.utc
    )
    assert row["worked_minutes"] == 605
    assert row["attendance_status"] == "complete"

    report = client.get(
        "/api/attendance/reports/daily.xlsx?day=2026-08-17&lang=en",
        headers=auth_headers,
    )
    assert report.status_code == 200, report.text
    assert report.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attendance_daily_2026-08-17.xlsx" in report.headers["content-disposition"]
    sheet = load_workbook(BytesIO(report.content), data_only=False).active
    assert sheet["A1"].value == "Daily attendance report"
    assert [sheet.cell(6, column).value for column in range(1, 8)] == [
        "No.",
        "Employee ID",
        "Employee",
        "Status",
        "Arrival (first scan)",
        "Departure (last scan)",
        "Time between scans",
    ]
    assert [sheet.cell(7, column).value for column in range(1, 8)] == [
        1,
        "735",
        "Test Person",
        "Arrival and departure",
        "08:15",
        "18:20",
        "10:05",
    ]


def test_daily_attendance_merges_the_same_employee_across_turnstiles(client, auth_headers):
    first_device = snapshot(
        person("735", "Test Person"),
        device_key="turnstile-1",
        source_host="10.100.50.73",
    )
    second_device = snapshot(
        person("735", "Test Person"),
        device_key="turnstile-2",
        source_host="10.100.50.31",
    )
    for payload in (first_device, second_device):
        response = client.post(
            "/api/attendance/integration/people",
            headers=INTEGRATION_HEADERS,
            json=payload,
        )
        assert response.status_code == 200, response.text

    for payload, event_uid, occurred_at in (
        (first_device, "lane-1-entry", "2026-08-17T08:10:00+05:00"),
        (second_device, "lane-2-exit", "2026-08-17T18:25:00+05:00"),
    ):
        response = client.post(
            "/api/attendance/integration/events",
            headers=INTEGRATION_HEADERS,
            json={
                "device": payload["device"],
                "events": [{
                    "event_uid": event_uid,
                    "external_person_id": "735",
                    "occurred_at": occurred_at,
                    "direction": "unknown",
                }],
            },
        )
        assert response.status_code == 200, response.text

    overview = client.get("/api/attendance/overview?day=2026-08-17", headers=auth_headers)
    assert overview.status_code == 200, overview.text
    body = overview.json()
    assert len(body["devices"]) == 2
    assert body["summary"]["total_people"] == 1
    assert body["summary"]["used_today"] == 1
    assert body["summary"]["events_today"] == 2
    assert body["pagination"]["total"] == 1
    assert len(body["people"]) == 1
    assert body["people"][0]["external_person_id"] == "735"
    assert body["people"][0]["worked_minutes"] == 615
    assert body["people"][0]["attendance_status"] == "complete"


def test_near_simultaneous_lane_duplicates_do_not_create_a_departure(client, auth_headers):
    payload = snapshot(person("735", "Test Person"))
    client.post(
        "/api/attendance/integration/people",
        headers=INTEGRATION_HEADERS,
        json=payload,
    )
    imported = client.post(
        "/api/attendance/integration/events",
        headers=INTEGRATION_HEADERS,
        json={
            "device": payload["device"],
            "events": [
                {
                    "event_uid": "duplicate-lane-a",
                    "external_person_id": "735",
                    "occurred_at": "2026-08-17T08:15:00+05:00",
                },
                {
                    "event_uid": "duplicate-lane-b",
                    "external_person_id": "735",
                    "occurred_at": "2026-08-17T08:15:30+05:00",
                },
            ],
        },
    )
    assert imported.status_code == 200, imported.text

    row = client.get(
        "/api/attendance/overview?day=2026-08-17",
        headers=auth_headers,
    ).json()["people"][0]
    assert row["event_count"] == 2
    assert row["departure_at"] is None
    assert row["worked_minutes"] is None
    assert row["attendance_status"] == "single_scan"


def test_full_snapshot_marks_removed_device_profile_without_deleting_it(client):
    client.post(
        "/api/attendance/integration/people",
        headers=INTEGRATION_HEADERS,
        json=snapshot(person("735", "One"), person("371", "Two")),
    )
    client.post(
        "/api/attendance/integration/people",
        headers=INTEGRATION_HEADERS,
        json=snapshot(person("735", "One")),
    )
    with TestSessionLocal() as db:
        removed = db.query(AttendancePerson).filter(AttendancePerson.external_person_id == "371").one()
        assert removed.present_on_device is False


def test_photo_is_private_and_image_validated(client, auth_headers):
    client.post(
        "/api/attendance/integration/people",
        headers=INTEGRATION_HEADERS,
        json=snapshot(person("735", "Photo Person")),
    )
    image = Image.new("RGB", (12, 12), (40, 50, 60))
    content = BytesIO()
    image.save(content, format="PNG")
    uploaded = client.post(
        "/api/attendance/integration/photos/main-turnstile/735",
        headers={**INTEGRATION_HEADERS, "Content-Type": "image/png"},
        content=content.getvalue(),
    )
    assert uploaded.status_code == 200, uploaded.text
    with TestSessionLocal() as db:
        person_id = db.query(AttendancePerson.id).filter(AttendancePerson.external_person_id == "735").scalar()
    assert client.get(f"/api/attendance/people/{person_id}/photo").status_code == 401
    photo = client.get(f"/api/attendance/people/{person_id}/photo", headers=auth_headers)
    assert photo.status_code == 200
    assert photo.headers["content-type"].startswith("image/webp")


def test_eco_admin_can_enroll_dahua_device_without_storing_plaintext_credentials(client):
    login = client.post(
        "/api/auth/login-json",
        json={
            "email": "admin@example.com",
            "password": "test-admin-password-123!",
            "factory_code": "ECO",
        },
    )
    assert login.status_code == 200, login.text
    created = client.post(
        "/api/attendance/devices",
        json={
            "device_key": "ect-dahua-main",
            "name": "ECT main entrance",
            "vendor": "Dahua",
            "source_host": "https://10.100.60.20",
            "certificate_sha256": "ab" * 32,
        },
    )
    assert created.status_code == 201, created.text
    device_payload = created.json()
    connector_token = device_payload["connector_token"]
    assert connector_token

    with TestSessionLocal() as db:
        device = db.get(AttendanceDevice, device_payload["id"])
        assert device.factory_code == "ECO"
        assert device.vendor == "Dahua"
        assert device.connector_token_hash != connector_token
        assert device.source_host == "https://10.100.60.20"

    imported = client.post(
        "/api/attendance/integration/people",
        headers={"X-Attendance-Token": connector_token},
        json=snapshot(
            person("ECT-001", "Eco Cotton Worker"),
            device_key="ect-dahua-main",
            source_host="https://10.100.60.20",
        ) | {"device": {
            **snapshot(device_key="ect-dahua-main", source_host="https://10.100.60.20")["device"],
            "vendor": "Dahua",
        }},
    )
    assert imported.status_code == 200, imported.text
    overview = client.get("/api/attendance/overview?day=2026-08-17")
    assert overview.status_code == 200, overview.text
    assert overview.json()["summary"]["total_people"] == 1
    assert overview.json()["devices"][0]["managed"] is True

    disabled = client.patch(
        f"/api/attendance/devices/{device_payload['id']}",
        json={"sync_enabled": False},
    )
    assert disabled.status_code == 200, disabled.text
    blocked = client.post(
        "/api/attendance/integration/people",
        headers={"X-Attendance-Token": connector_token},
        json=snapshot(person("ECT-001", "Eco Cotton Worker"), device_key="ect-dahua-main"),
    )
    assert blocked.status_code == 403
